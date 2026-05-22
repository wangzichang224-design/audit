from __future__ import annotations

import json
import re
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from auditpaper_agent.config import Settings, get_settings


@dataclass(frozen=True)
class AgentMaterialSelection:
    selected_paths: dict[str, str] = field(default_factory=dict)
    confidence: float = 0.0
    reason: str = ""
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class SchemaMappingResult:
    path: Path
    target_schema: str
    mapping: dict[str, str] = field(default_factory=dict)
    confidence: float = 0.0
    reason: str = ""
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class ProviderStatus:
    provider: str
    model: str
    base_url: str
    configured: bool
    ok: bool
    purpose: str
    error: str = ""


@dataclass(frozen=True)
class ProviderHealth:
    reasoning: ProviderStatus
    vision: ProviderStatus
    checked_live: bool = False


class AgentReasoningProvider:
    """OpenAI-compatible agent layer for material selection and schema mapping.

    The provider may inspect file metadata, headers, and a few sample cells. It
    must not mutate workbooks or make audit decisions that bypass the harness.
    """

    def __init__(self, settings: Settings | None = None) -> None:
        self.settings = settings or get_settings()

    def select_cash_materials(self, root: Path, material_sets: Iterable[Any]) -> AgentMaterialSelection:
        cfg = self.settings.reasoning
        cfg.require_api_key("Agent material selection")
        candidates = [_material_set_payload(item) for item in material_sets]
        content = _post_openai_chat(
            base_url=cfg.base_url,
            api_key=cfg.api_key,
            model=cfg.model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are AuditPaper-Agent's material resolver. Choose one coherent C/cash audit material package. "
                        "Do not mix files from different client projects. Return strict JSON only."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "root": str(root),
                            "task": "select C cash materials",
                            "required_keys": ["materials_dir", "trial_balance", "journal", "bank_statement", "template"],
                            "rules": [
                                "Prefer one c底稿资料 folder or one project root.",
                                "Reject AP/AR/customer/supplier subledgers as bank journal.",
                                "A bank journal must have date, summary/description, account/cash account, and debit/credit/amount columns.",
                            ],
                            "candidates": candidates[:8],
                            "response_schema": {
                                "selected_paths": {
                                    "materials_dir": "string",
                                    "trial_balance": "string",
                                    "journal": "string",
                                    "bank_statement": "string or empty",
                                    "template": "string",
                                },
                                "confidence": "0..1 number",
                                "reason": "short Chinese explanation",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        )
        data = _extract_json_object(content)
        selected = data.get("selected_paths") if isinstance(data.get("selected_paths"), dict) else {}
        return AgentMaterialSelection(
            selected_paths={str(k): str(v) for k, v in selected.items() if v not in (None, "")},
            confidence=_coerce_confidence(data.get("confidence")),
            reason=str(data.get("reason") or ""),
            raw=data,
        )

    def map_table_schema(self, path: Path, target_schema: str) -> SchemaMappingResult:
        cfg = self.settings.reasoning
        cfg.require_api_key("Agent schema mapping")
        profile = _table_profile(path)
        content = _post_openai_chat(
            base_url=cfg.base_url,
            api_key=cfg.api_key,
            model=cfg.model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You map client spreadsheet headers to AuditPaper-Agent canonical schemas. "
                        "Return strict JSON only and never invent missing columns."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "path": str(path),
                            "target_schema": target_schema,
                            "canonical_fields": _canonical_fields(target_schema),
                            "table_profile": profile,
                            "response_schema": {
                                "mapping": {"canonical_field": "source header"},
                                "confidence": "0..1 number",
                                "reason": "short Chinese explanation",
                            },
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        )
        data = _extract_json_object(content)
        mapping = data.get("mapping") if isinstance(data.get("mapping"), dict) else {}
        return SchemaMappingResult(
            path=path,
            target_schema=target_schema,
            mapping={str(k): str(v) for k, v in mapping.items() if v not in (None, "")},
            confidence=_coerce_confidence(data.get("confidence")),
            reason=str(data.get("reason") or ""),
            raw=data,
        )


def map_table_schema_with_agent(path: str | Path, target_schema: str) -> SchemaMappingResult:
    return AgentReasoningProvider().map_table_schema(Path(path), target_schema)


def check_provider_health(live: bool = False, settings: Settings | None = None) -> ProviderHealth:
    cfg = settings or get_settings()
    reasoning_ok, reasoning_error = _check_chat_provider(cfg.reasoning, live=live) if cfg.reasoning.is_configured else (False, "")
    vision_ok, vision_error = _check_chat_provider(cfg.vision, live=live) if cfg.vision.is_configured else (False, "")
    return ProviderHealth(
        reasoning=ProviderStatus(
            provider=cfg.reasoning.provider,
            model=cfg.reasoning.model,
            base_url=cfg.reasoning.base_url,
            configured=cfg.reasoning.is_configured,
            ok=reasoning_ok if live else cfg.reasoning.is_configured,
            purpose="material/schema reasoning and finding wording",
            error=reasoning_error,
        ),
        vision=ProviderStatus(
            provider=cfg.vision.provider,
            model=cfg.vision.model,
            base_url=cfg.vision.base_url,
            configured=cfg.vision.is_configured,
            ok=vision_ok if live else cfg.vision.is_configured,
            purpose="OCR for scanned/image confirmations",
            error=vision_error,
        ),
        checked_live=live,
    )


def _check_chat_provider(provider_cfg, live: bool) -> tuple[bool, str]:
    if not provider_cfg.api_key:
        return False, "not configured"
    if not live:
        return True, ""
    try:
        content = _post_openai_chat(
            base_url=provider_cfg.base_url,
            api_key=provider_cfg.api_key,
            model=provider_cfg.model,
            messages=[
                {"role": "system", "content": "Return strict JSON only."},
                {"role": "user", "content": '{"ok": true}'},
            ],
            timeout=20,
        )
        _extract_json_object(content)
        return True, ""
    except Exception as exc:
        return False, str(exc)


def _material_set_payload(item: Any) -> dict[str, Any]:
    discovery = item.discovery
    return {
        "root": str(item.root),
        "score": item.score,
        "confidence": item.confidence,
        "reasons": item.reasons,
        "trial_balance": _file_profile(discovery.trial_balance),
        "journal": _file_profile(discovery.journal),
        "bank_statement": str(discovery.bank_statement or ""),
        "template": _file_profile(discovery.template),
        "missing_required": discovery.missing_required,
    }


def _file_profile(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    payload: dict[str, Any] = {"path": str(path), "name": path.name, "suffix": path.suffix}
    if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".csv"}:
        payload.update(_table_profile(path))
    return payload


def _table_profile(path: Path) -> dict[str, Any]:
    try:
        if path.suffix.lower() == ".csv":
            df = pd.read_csv(path, encoding="utf-8-sig", nrows=5)
            return {"sheet": "", "headers": [str(c) for c in df.columns], "sample_rows": _sample_rows(df)}

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            best_sheet = ""
            best_headers: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
                    values = [str(v).strip() for v in row if v not in (None, "")]
                    if len(values) > len(best_headers):
                        best_sheet = ws.title
                        best_headers = values
                if len(best_headers) >= 3:
                    break
        finally:
            wb.close()
        return {"sheet": best_sheet, "headers": best_headers[:40], "sample_rows": []}
    except Exception as exc:
        return {"error": str(exc)}


def _sample_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in df.head(3).to_dict(orient="records"):
        rows.append({str(k): _json_cell(v) for k, v in row.items()})
    return rows


def _canonical_fields(target_schema: str) -> list[str]:
    schemas = {
        "cash_journal": ["date", "description", "account_code", "account_name", "debit", "credit", "balance", "counterparty", "txn_id"],
        "trial_balance": ["account_code", "account_name", "ending_debit", "ending_credit", "ending_balance", "prior_year"],
        "investment": ["account_code", "account_name", "ending_balance", "fair_value", "classification", "issuer", "product_name"],
        "expense": ["date", "voucher_id", "account_code", "account_name", "description", "debit", "credit", "amount"],
    }
    return schemas.get(target_schema, [])


def _post_openai_chat(
    base_url: str,
    api_key: str,
    model: str,
    messages: list[dict[str, str]],
    timeout: int = 60,
) -> str:
    payload = {"model": model, "temperature": 0, "messages": messages}
    request = urllib.request.Request(
        f"{base_url.rstrip('/')}/chat/completions",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        data = json.loads(response.read().decode("utf-8"))
    try:
        return str(data["choices"][0]["message"]["content"])
    except (KeyError, IndexError, TypeError) as exc:
        raise RuntimeError(f"Unexpected provider response: {data}") from exc


def _extract_json_object(content: str) -> dict[str, Any]:
    text = content.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"```$", "", text).strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if not match:
            raise ValueError(f"Provider did not return JSON: {content[:200]!r}")
        data = json.loads(match.group(0))
    if not isinstance(data, dict):
        raise ValueError("Provider response must be a JSON object.")
    return data


def _coerce_confidence(value: Any) -> float:
    try:
        return max(0.0, min(float(value), 1.0))
    except Exception:
        return 0.0


def _json_cell(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value
