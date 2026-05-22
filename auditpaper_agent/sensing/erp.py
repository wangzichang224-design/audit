from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from auditpaper_agent.contracts import (
    BankJournalRow,
    CaseMetadata,
    DocumentType,
    ErpFieldMapping,
    ErpMappingManifest,
    ErpTableMapping,
    GeneralLedgerRow,
    MasterDataRecord,
    SourceDocument,
    SourceRef,
    StandardizedErpAuditPackage,
    TrialBalanceRow,
)
from auditpaper_agent.utils import parse_float, safe_filename, sha256_file, write_json


ERP_TABLE_TYPES: tuple[DocumentType, ...] = (
    "trial_balance",
    "general_ledger",
    "customers",
    "suppliers",
    "inventory",
    "fixed_assets",
    "sales",
    "purchase",
    "bank_statement",
)


TYPE_TERMS: dict[DocumentType, tuple[str, ...]] = {
    "trial_balance": ("试算", "科目余额", "余额表", "trial", "balance", "s_alr_87012301", "gl_bal"),
    "general_ledger": ("序时", "总账", "明细账", "凭证", "journal", "ledger", "fbl3n", "gl_line", "bkpf", "bseg"),
    "customers": ("客户", "customer", "kna1", "应收客户"),
    "suppliers": ("供应商", "vendor", "supplier", "lfa1", "应付供应商"),
    "inventory": ("存货", "库存", "物料", "inventory", "material", "mb52", "mard"),
    "fixed_assets": ("固定资产", "资产卡片", "asset", "anla", "折旧"),
    "sales": ("销售", "收入", "sales", "revenue", "vf05", "billing"),
    "purchase": ("采购", "purchase", "procurement", "me2n", "po_", "采购订单"),
    "bank_statement": ("银行", "对账单", "bank", "statement"),
    "unknown": (),
    "bank_journal": (),
    "bank_confirmation": (),
}


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "account_code": ("科目编码", "科目代码", "总账科目", "会计科目", "account", "g/l acct", "g/l account", "saknr"),
    "account_name": ("科目名称", "账户名称", "科目描述", "account name", "g/l acct long text", "txt50"),
    "ending_debit": ("期末借方", "借方余额", "ending debit", "debit balance"),
    "ending_credit": ("期末贷方", "贷方余额", "ending credit", "credit balance"),
    "ending_balance": ("期末余额", "余额", "本币余额", "ending balance", "balance", "hsl", "local currency amount"),
    "prior_year": ("期初余额", "年初余额", "上年末审定数", "prior year", "opening balance"),
    "direction": ("余额方向", "借贷方向", "debit/credit", "dr/cr", "shkzg"),
    "posting_date": ("过账日期", "凭证日期", "记账日期", "业务日期", "date", "posting date", "budat"),
    "voucher_id": ("凭证号", "凭证编号", "会计凭证", "document number", "belnr", "voucher"),
    "line_no": ("行项目", "行号", "line item", "buzei"),
    "company_code": ("公司代码", "公司", "company code", "bukrs"),
    "description": ("摘要", "文本", "项目文本", "说明", "description", "text", "sgtxt"),
    "debit": ("借方金额", "借方", "debit", "debit amount"),
    "credit": ("贷方金额", "贷方", "credit", "credit amount"),
    "amount": ("金额", "本币金额", "发生额", "amount", "dmbtr", "wrbtr"),
    "currency": ("币种", "货币", "currency", "waers"),
    "counterparty": ("对方单位", "往来单位", "counterparty", "trading partner"),
    "customer_id": ("客户编码", "客户编号", "客户", "customer", "kunnr"),
    "customer_name": ("客户名称", "customer name", "name1"),
    "supplier_id": ("供应商编码", "供应商编号", "供应商", "vendor", "lifnr"),
    "supplier_name": ("供应商名称", "vendor name", "supplier name", "name1"),
    "material_id": ("物料编码", "存货编码", "产品编码", "material", "matnr", "sku"),
    "material_name": ("物料名称", "存货名称", "产品名称", "material description", "maktx"),
    "asset_id": ("资产编码", "固定资产编号", "asset", "anln1"),
    "asset_name": ("资产名称", "fixed asset name", "asset description"),
    "quantity": ("数量", "qty", "quantity", "menge"),
    "unit_cost": ("单位成本", "单价", "unit cost", "price"),
    "cost": ("原值", "成本", "金额", "cost", "amount"),
    "sales_order": ("销售订单", "订单号", "sales order", "vbeln"),
    "purchase_order": ("采购订单", "po", "purchase order", "ebeln"),
}


REQUIRED_FIELDS: dict[DocumentType, tuple[str, ...]] = {
    "trial_balance": ("account_code", "account_name"),
    "general_ledger": ("account_code", "account_name"),
    "customers": ("customer_id", "customer_name"),
    "suppliers": ("supplier_id", "supplier_name"),
    "inventory": ("material_id", "material_name"),
    "fixed_assets": ("asset_id", "asset_name"),
    "sales": ("amount",),
    "purchase": ("amount",),
    "bank_statement": ("posting_date",),
}


@dataclass(frozen=True)
class ErpImportResult:
    success: bool
    case_dir: Path
    manifest_path: Path
    package_path: Path | None = None
    package: StandardizedErpAuditPackage | None = None
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _TableData:
    path: Path
    df: pd.DataFrame
    sheet_name: str = ""
    header_row: int = 1


def diagnose_erp_export(
    export_dir: str | Path,
    provider: str = "auto",
    confirmed: bool = False,
) -> ErpMappingManifest:
    root = _clean_path(export_dir)
    if not root.is_dir():
        return ErpMappingManifest(
            provider=_provider(provider),
            root_path=str(root),
            confirmed=False,
            generated_at=_now(),
            tables=[],
            blocking_issues=[f"ERP export folder does not exist: {root}"],
        )

    tables: list[ErpTableMapping] = []
    for table in _read_candidate_tables(root):
        source_type = _classify_table(table, provider)
        if source_type == "unknown":
            continue
        mapping = _build_table_mapping(table, source_type)
        tables.append(mapping)

    blocking = _blocking_issues(tables)
    return ErpMappingManifest(
        provider=_provider(provider),
        root_path=str(root),
        confirmed=confirmed,
        generated_at=_now(),
        tables=tables,
        blocking_issues=blocking,
    )


def write_erp_mapping_manifest(
    export_dir: str | Path,
    output_dir: str | Path,
    provider: str = "auto",
    confirmed: bool = False,
) -> Path:
    manifest = diagnose_erp_export(export_dir, provider=provider, confirmed=confirmed)
    path = Path(output_dir) / "mapping_manifest.json"
    write_json(path, manifest)
    return path


def import_erp_export(
    export_dir: str | Path,
    case_dir: str | Path,
    provider: str = "auto",
    confirm_mapping: bool = False,
) -> ErpImportResult:
    output = Path(case_dir)
    output.mkdir(parents=True, exist_ok=True)
    manifest = diagnose_erp_export(export_dir, provider=provider, confirmed=confirm_mapping)
    manifest_path = output / "mapping_manifest.json"
    write_json(manifest_path, manifest)

    if not manifest.confirmed:
        return ErpImportResult(
            success=False,
            case_dir=output,
            manifest_path=manifest_path,
            errors=["字段映射尚未确认。请复核 mapping_manifest.json 后使用 confirm_mapping=True 或 CLI --confirm-mapping。"],
        )
    if manifest.blocking_issues:
        return ErpImportResult(success=False, case_dir=output, manifest_path=manifest_path, errors=manifest.blocking_issues)

    package = _standardize_export(_clean_path(export_dir), manifest)
    package_path = output / "erp_package.json"
    write_json(package_path, package)
    return ErpImportResult(
        success=True,
        case_dir=output,
        manifest_path=manifest_path,
        package_path=package_path,
        package=package,
    )


def load_erp_package(case_dir: str | Path) -> StandardizedErpAuditPackage:
    data = json.loads((Path(case_dir) / "erp_package.json").read_text(encoding="utf-8"))
    return StandardizedErpAuditPackage.model_validate(data)


def _standardize_export(root: Path, manifest: ErpMappingManifest) -> StandardizedErpAuditPackage:
    table_data = {str(table.path): _read_table(Path(table.path), header_row=table.header_row, sheet_name=table.sheet_name) for table in manifest.tables}
    meta = _infer_meta(root, table_data.values())
    docs = [_source_document(Path(table.path), table.source_type) for table in manifest.tables]
    trial_balance: list[TrialBalanceRow] = []
    general_ledger: list[GeneralLedgerRow] = []
    customers: list[MasterDataRecord] = []
    suppliers: list[MasterDataRecord] = []
    inventory: list[MasterDataRecord] = []
    fixed_assets: list[MasterDataRecord] = []
    sales: list[MasterDataRecord] = []
    purchase: list[MasterDataRecord] = []

    for table in manifest.tables:
        data = table_data[str(table.path)]
        mapping = _mapping_dict(table)
        if table.source_type == "trial_balance":
            trial_balance.extend(_parse_trial_balance(data, mapping))
        elif table.source_type == "general_ledger":
            general_ledger.extend(_parse_general_ledger(data, mapping))
        elif table.source_type == "customers":
            customers.extend(_parse_master(data, mapping, "customer", id_field="customer_id", name_field="customer_name"))
        elif table.source_type == "suppliers":
            suppliers.extend(_parse_master(data, mapping, "supplier", id_field="supplier_id", name_field="supplier_name"))
        elif table.source_type == "inventory":
            inventory.extend(_parse_master(data, mapping, "inventory", id_field="material_id", name_field="material_name"))
        elif table.source_type == "fixed_assets":
            fixed_assets.extend(_parse_master(data, mapping, "fixed_asset", id_field="asset_id", name_field="asset_name"))
        elif table.source_type == "sales":
            sales.extend(_parse_master(data, mapping, "sales", id_field="sales_order", name_field="customer_name"))
        elif table.source_type == "purchase":
            purchase.extend(_parse_master(data, mapping, "purchase", id_field="purchase_order", name_field="supplier_name"))

    bank_journal = [_gl_to_bank_journal(row) for row in general_ledger if _is_cash_account(row.account_code, row.account_name)]
    return StandardizedErpAuditPackage(
        meta=meta,
        provider=manifest.provider,
        source_documents=docs,
        mapping_confirmed=manifest.confirmed,
        trial_balance=trial_balance,
        general_ledger=general_ledger,
        bank_journal=bank_journal,
        customers=customers,
        suppliers=suppliers,
        inventory=inventory,
        fixed_assets=fixed_assets,
        sales=sales,
        purchase=purchase,
    )


def _read_candidate_tables(root: Path) -> list[_TableData]:
    tables: list[_TableData] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".xlsm", ".xls", ".csv"}:
            continue
        try:
            tables.append(_read_table(path))
        except Exception:
            continue
    return tables


def _read_table(path: Path, header_row: int | None = None, sheet_name: str = "") -> _TableData:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(path, encoding="utf-8-sig", header=None, dtype=object)
        header_idx = (header_row - 1) if header_row else _best_header_row(raw)
        df = pd.read_csv(path, encoding="utf-8-sig", header=header_idx, dtype=object)
        return _TableData(path=path, df=_clean_columns(df), header_row=header_idx + 1)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        chosen_sheet = sheet_name if sheet_name in wb.sheetnames else _best_sheet(wb)
        ws = wb[chosen_sheet]
        preview = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 24), values_only=True):
            preview.append([cell for cell in row])
    finally:
        wb.close()

    preview_df = pd.DataFrame(preview)
    header_idx = (header_row - 1) if header_row else _best_header_row(preview_df)
    df = pd.read_excel(path, sheet_name=chosen_sheet, header=header_idx, dtype=object)
    return _TableData(path=path, df=_clean_columns(df), sheet_name=str(chosen_sheet), header_row=header_idx + 1)


def _best_sheet(workbook) -> str:
    best = workbook.sheetnames[0]
    best_score = -1
    for ws in workbook.worksheets:
        values: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
            values.extend(str(v) for v in row if v not in (None, ""))
        score = _header_score(values)
        if score > best_score:
            best = ws.title
            best_score = score
    return str(best)


def _best_header_row(raw: pd.DataFrame) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in raw.head(24).iterrows():
        values = [str(value) for value in row.tolist() if value not in (None, "") and str(value) != "nan"]
        score = _header_score(values)
        if score > best_score:
            best_idx = int(idx)
            best_score = score
    return best_idx


def _header_score(values: Iterable[str]) -> int:
    text = " ".join(values).lower()
    score = 0
    for aliases in FIELD_ALIASES.values():
        for alias in aliases:
            if _norm(alias) in _norm(text):
                score += 1
                break
    return score


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    clean = df.copy()
    seen: dict[str, int] = {}
    columns: list[str] = []
    for idx, column in enumerate(clean.columns):
        name = str(column).strip()
        if not name or name.lower().startswith("unnamed"):
            name = f"column_{idx + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        columns.append(name)
    clean.columns = columns
    clean = clean.dropna(how="all")
    return clean


def _classify_table(table: _TableData, provider: str) -> DocumentType:
    file_text = f"{table.path.name} {table.path.parent.name}".lower()
    header_text = " ".join(str(c) for c in table.df.columns).lower()
    scored: list[tuple[int, DocumentType]] = []
    for table_type in ERP_TABLE_TYPES:
        score = 0
        score += sum(4 for token in TYPE_TERMS[table_type] if token.lower() in file_text)
        score += sum(2 for token in TYPE_TERMS[table_type] if token.lower() in header_text)
        required = REQUIRED_FIELDS.get(table_type, ())
        matched = _match_fields(table.df.columns, required)
        score += 3 * len([field for field in required if matched.get(field)])
        if table_type == "trial_balance" and _match_fields(table.df.columns, ("ending_balance", "ending_debit", "ending_credit")):
            score += 4
        if table_type == "general_ledger" and _match_fields(table.df.columns, ("posting_date", "voucher_id", "debit", "credit", "amount")):
            score += 4
        scored.append((score, table_type))
    scored.sort(key=lambda item: (-item[0], item[1]))
    return scored[0][1] if scored and scored[0][0] >= 5 else "unknown"


def _build_table_mapping(table: _TableData, source_type: DocumentType) -> ErpTableMapping:
    canonical_fields = _canonical_fields_for_type(source_type)
    matched = _match_fields(table.df.columns, canonical_fields)
    required = REQUIRED_FIELDS.get(source_type, ())
    missing = [field for field in required if not matched.get(field)]
    mappings = [
        ErpFieldMapping(
            canonical_field=field,
            source_header=matched.get(field, ""),
            confidence=1.0 if matched.get(field) else 0.0,
            required=field in required,
            notes="" if matched.get(field) else "未识别，需人工补充或确认不适用",
        )
        for field in canonical_fields
    ]
    debit_total = _sum_field(table.df, matched.get("debit"))
    credit_total = _sum_field(table.df, matched.get("credit"))
    balance_total = _sum_field(table.df, matched.get("ending_balance") or matched.get("amount"))
    matched_count = len([item for item in mappings if item.source_header])
    confidence = round(min(1.0, matched_count / max(len(required) + 2, 1)), 3)
    return ErpTableMapping(
        source_type=source_type,
        document_id=table.path.stem,
        path=str(table.path),
        sheet_name=table.sheet_name,
        row_count=len(table.df),
        header_row=table.header_row,
        matched_fields=mappings,
        missing_required=missing,
        confidence=confidence,
        sample_rows=_sample_rows(table.df),
        debit_total=round(debit_total, 2),
        credit_total=round(credit_total, 2),
        balance_total=round(balance_total, 2),
    )


def _canonical_fields_for_type(source_type: DocumentType) -> tuple[str, ...]:
    if source_type == "trial_balance":
        return ("account_code", "account_name", "ending_debit", "ending_credit", "ending_balance", "direction", "prior_year", "currency", "company_code")
    if source_type == "general_ledger":
        return (
            "posting_date",
            "voucher_id",
            "line_no",
            "company_code",
            "account_code",
            "account_name",
            "description",
            "debit",
            "credit",
            "amount",
            "direction",
            "currency",
            "counterparty",
            "customer_id",
            "supplier_id",
            "material_id",
            "asset_id",
        )
    if source_type == "customers":
        return ("customer_id", "customer_name", "amount", "currency")
    if source_type == "suppliers":
        return ("supplier_id", "supplier_name", "amount", "currency")
    if source_type == "inventory":
        return ("material_id", "material_name", "quantity", "unit_cost", "cost", "amount", "currency")
    if source_type == "fixed_assets":
        return ("asset_id", "asset_name", "cost", "amount", "currency")
    if source_type == "sales":
        return ("posting_date", "sales_order", "customer_id", "customer_name", "amount", "currency")
    if source_type == "purchase":
        return ("posting_date", "purchase_order", "supplier_id", "supplier_name", "amount", "currency")
    return tuple(FIELD_ALIASES)


def _match_fields(headers: Iterable[Any], fields: Iterable[str]) -> dict[str, str]:
    normalized = {_norm(str(header)): str(header) for header in headers}
    result: dict[str, str] = {}
    for field in fields:
        aliases = FIELD_ALIASES.get(field, (field,))
        for alias in aliases:
            needle = _norm(alias)
            for norm_header, original in normalized.items():
                if needle == norm_header or needle in norm_header:
                    result[field] = original
                    break
            if field in result:
                break
    return result


def _blocking_issues(tables: list[ErpTableMapping]) -> list[str]:
    issues: list[str] = []
    by_type = {table.source_type: table for table in tables}
    for required_type, label in {"trial_balance": "试算平衡表/TB", "general_ledger": "总账/序时账"}.items():
        table = by_type.get(required_type)
        if table is None:
            issues.append(f"缺少必要 ERP 导出：{label}")
        elif table.missing_required:
            issues.append(f"{label} 缺少关键字段：{', '.join(table.missing_required)}")
    return issues


def _mapping_dict(table: ErpTableMapping) -> dict[str, str]:
    return {field.canonical_field: field.source_header for field in table.matched_fields if field.source_header}


def _parse_trial_balance(table: _TableData, mapping: dict[str, str]) -> list[TrialBalanceRow]:
    rows: list[TrialBalanceRow] = []
    for idx, raw in table.df.iterrows():
        account_name = _value(raw, mapping, "account_name")
        account_code = _string(_value(raw, mapping, "account_code"))
        if not account_code and not account_name:
            continue
        ending_debit = _float(_value(raw, mapping, "ending_debit"))
        ending_credit = _float(_value(raw, mapping, "ending_credit"))
        if not ending_debit and not ending_credit:
            balance = _float(_value(raw, mapping, "ending_balance"))
            direction = _string(_value(raw, mapping, "direction"))
            if _is_credit_direction(direction):
                ending_credit = abs(balance)
            else:
                ending_debit = balance
        rows.append(
            TrialBalanceRow(
                account_code=account_code,
                account_name=_string(account_name),
                ending_debit=ending_debit,
                ending_credit=ending_credit,
                prior_year=_float(_value(raw, mapping, "prior_year")),
                source=_source_ref(table.path, "trial_balance", table.sheet_name, int(idx) + table.header_row + 1),
            )
        )
    return rows


def _parse_general_ledger(table: _TableData, mapping: dict[str, str]) -> list[GeneralLedgerRow]:
    rows: list[GeneralLedgerRow] = []
    for idx, raw in table.df.iterrows():
        account_name = _value(raw, mapping, "account_name")
        account_code = _string(_value(raw, mapping, "account_code"))
        if not account_code and not account_name:
            continue
        debit = _float(_value(raw, mapping, "debit"))
        credit = _float(_value(raw, mapping, "credit"))
        if not debit and not credit:
            amount = _float(_value(raw, mapping, "amount"))
            direction = _string(_value(raw, mapping, "direction"))
            if _is_credit_direction(direction) or amount < 0:
                credit = abs(amount)
            else:
                debit = abs(amount)
        rows.append(
            GeneralLedgerRow(
                posting_date=_date(_value(raw, mapping, "posting_date")),
                voucher_id=_string(_value(raw, mapping, "voucher_id")),
                line_no=_string(_value(raw, mapping, "line_no")),
                company_code=_string(_value(raw, mapping, "company_code")),
                account_code=account_code,
                account_name=_string(account_name),
                description=_string(_value(raw, mapping, "description")),
                debit=debit,
                credit=credit,
                currency=_string(_value(raw, mapping, "currency")) or "CNY",
                counterparty=_string(_value(raw, mapping, "counterparty")),
                customer_id=_string(_value(raw, mapping, "customer_id")),
                supplier_id=_string(_value(raw, mapping, "supplier_id")),
                material_id=_string(_value(raw, mapping, "material_id")),
                asset_id=_string(_value(raw, mapping, "asset_id")),
                source=_source_ref(table.path, "general_ledger", table.sheet_name, int(idx) + table.header_row + 1),
            )
        )
    return rows


def _parse_master(
    table: _TableData,
    mapping: dict[str, str],
    record_type: str,
    id_field: str,
    name_field: str,
) -> list[MasterDataRecord]:
    rows: list[MasterDataRecord] = []
    for idx, raw in table.df.iterrows():
        record_id = _string(_value(raw, mapping, id_field))
        name = _string(_value(raw, mapping, name_field))
        if not record_id and not name:
            continue
        amount = _float(_value(raw, mapping, "amount")) or _float(_value(raw, mapping, "cost"))
        attrs = {str(k): _json_cell(v) for k, v in raw.items() if v not in (None, "") and not pd.isna(v)}
        rows.append(
            MasterDataRecord(
                record_type=record_type,
                record_id=record_id,
                name=name,
                amount=amount,
                currency=_string(_value(raw, mapping, "currency")) or "CNY",
                date_value=_date(_value(raw, mapping, "posting_date")),
                attributes=attrs,
                source=_source_ref(table.path, _record_source_type(record_type), table.sheet_name, int(idx) + table.header_row + 1),
            )
        )
    return rows


def _gl_to_bank_journal(row: GeneralLedgerRow) -> BankJournalRow:
    return BankJournalRow(
        txn_date=row.posting_date or date.today(),
        bank_name="",
        bank_account="",
        currency=row.currency,
        description=row.description,
        debit=row.debit,
        credit=row.credit,
        balance=0.0,
        counterparty=row.counterparty,
        txn_id=row.voucher_id,
        source=row.source,
    )


def _infer_meta(root: Path, tables: Iterable[_TableData]) -> CaseMetadata:
    client_name = root.name
    period_end: date | None = None
    for table in tables:
        for column in table.df.columns:
            if _norm(str(column)) in {_norm("posting_date"), _norm("凭证日期"), _norm("过账日期")}:
                dates = [_date(value) for value in table.df[column].dropna().tolist()]
                dates = [value for value in dates if value is not None]
                if dates:
                    period_end = max(dates)
                    break
        if period_end:
            break
    return CaseMetadata(
        case_id=safe_filename(root.name),
        client_name=client_name,
        period_end=period_end,
        currency="CNY",
        gaap="企业会计准则",
    )


def _source_document(path: Path, document_type: DocumentType) -> SourceDocument:
    return SourceDocument(document_id=path.stem, path=str(path), document_type=document_type, sha256=sha256_file(path))


def _source_ref(path: Path, document_type: DocumentType, sheet: str = "", row: int | None = None) -> SourceRef:
    return SourceRef(
        document_id=path.stem,
        document_hash=sha256_file(path),
        file_path=str(path),
        sheet_name=sheet or None,
        row_number=row,
        locator=document_type,
    )


def _record_source_type(record_type: str) -> DocumentType:
    return {
        "customer": "customers",
        "supplier": "suppliers",
        "inventory": "inventory",
        "fixed_asset": "fixed_assets",
        "sales": "sales",
        "purchase": "purchase",
    }.get(record_type, "unknown")


def _sample_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in df.head(3).to_dict(orient="records"):
        rows.append({str(key): _json_cell(value) for key, value in row.items()})
    return rows


def _sum_field(df: pd.DataFrame, header: str | None) -> float:
    if not header or header not in df:
        return 0.0
    return sum(_float(value) for value in df[header].tolist())


def _value(row: Any, mapping: dict[str, str], field: str) -> Any:
    header = mapping.get(field)
    if not header or header not in row:
        return None
    value = row[header]
    if pd.isna(value):
        return None
    return value


def _float(value: Any) -> float:
    return parse_float(value, 0.0)


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_credit_direction(value: str) -> bool:
    text = str(value).strip().lower()
    return text in {"贷", "credit", "cr", "h"} or "贷" in text


def _is_cash_account(account_code: str, account_name: str) -> bool:
    return str(account_code).startswith(("1001", "1002", "1009")) or str(account_name).startswith(("库存现金", "银行存款", "其他货币资金"))


def _json_cell(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _norm(value: str) -> str:
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .replace("/", "")
        .replace(".", "")
        .replace("：", "")
        .replace(":", "")
    )


def _provider(value: str) -> str:
    normalized = (value or "auto").strip().lower()
    if normalized in {"sap", "yonyou", "auto"}:
        return normalized
    return "auto"


def _clean_path(value: str | Path) -> Path:
    text = str(value).strip().strip("\ufeff").strip()
    text = text.strip("`").strip().strip('"').strip("'").strip()
    if text.lower().startswith("file://"):
        text = text[7:]
    return Path(text).expanduser()


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")
