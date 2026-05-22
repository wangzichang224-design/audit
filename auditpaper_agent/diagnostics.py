from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

from auditpaper_agent.agent import check_provider_health
from auditpaper_agent.service import inspect_cash_materials, run_auto_cash_case
from auditpaper_agent.suite import inspect_workpaper_project, run_auto_workpaper_suite
from auditpaper_agent.utils import safe_filename


FILE_CLASS_TERMS: dict[str, tuple[str, ...]] = {
    "trial_balance": ("tb", "trial", "balance", "试算", "科目余额", "余额表"),
    "journal": ("journal", "ledger", "gl", "序时", "明细账", "凭证"),
    "bank": ("bank", "statement", "confirmation", "银行", "对账单", "回函", "函证"),
    "template": ("template", "workpaper", "底稿", "模板"),
    "master_data": ("master", "customer", "supplier", "product", "asset", "主数据", "客户", "供应商", "资产"),
    "sales_cycle": ("sales", "revenue", "销售", "收入"),
    "purchase_cycle": ("purchase", "procurement", "采购"),
}


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "account_code": ("account_code", "acct code", "account no", "科目编码", "科目代码", "科目编号"),
    "account_name": ("account_name", "acct name", "account title", "科目名称", "账户名称", "科目"),
    "ending_balance": ("ending_balance", "ending balance", "balance", "期末余额", "期末数", "余额"),
    "debit": ("debit", "dr", "借方金额", "本期借方", "借方发生额"),
    "credit": ("credit", "cr", "贷方金额", "本期贷方", "贷方发生额"),
    "date": ("date", "voucher_date", "posting date", "交易日期", "凭证日期", "日期"),
    "amount": ("amount", "value", "金额", "发生额", "本币金额"),
}


@dataclass(frozen=True)
class DiagnosticFile:
    path: Path
    category: str
    readable: bool
    sheet_name: str = ""
    headers: list[str] = field(default_factory=list)
    matched_columns: dict[str, str] = field(default_factory=dict)
    error: str = ""


@dataclass(frozen=True)
class DiagnosticCheck:
    name: str
    passed: bool
    detail: str
    confidence: float = 0.0


@dataclass(frozen=True)
class ClientPackageDiagnostics:
    input_path: Path
    mode: str
    can_generate: bool
    confidence: float
    project_dir: Path | None = None
    cash_materials_dir: Path | None = None
    missing_required: list[str] = field(default_factory=list)
    files: list[DiagnosticFile] = field(default_factory=list)
    checks: list[DiagnosticCheck] = field(default_factory=list)
    provider_summary: dict[str, str] = field(default_factory=dict)
    selected_cash_materials: dict[str, str] = field(default_factory=dict)
    cash_candidates: list[dict[str, Any]] = field(default_factory=list)
    agent_used: bool = False
    agent_reason: str = ""
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


@dataclass(frozen=True)
class StressCaseResult:
    case_dir: Path
    mode: str
    can_generate: bool
    success: bool
    generated_codes: list[str] = field(default_factory=list)
    missing_required: list[str] = field(default_factory=list)
    diagnostics_confidence: float = 0.0
    output_dir: Path | None = None
    error: str = ""


@dataclass(frozen=True)
class StressTestResult:
    fixtures_dir: Path
    focus: tuple[str, ...]
    case_count: int
    runnable_count: int
    success_count: int
    invalid_handled_count: int
    success_rate: float
    cases: list[StressCaseResult]
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


def diagnose_client_package(project_dir: str | Path) -> ClientPackageDiagnostics:
    input_path = _clean_path(project_dir)
    health = check_provider_health(live=False)
    provider_summary = {
        "default_ocr": "pdf-text/offline",
        "optional_ocr": "qwen-ocr, textin",
        "optional_reasoning": "DeepSeek wording and Agent material/schema reasoning",
        "api_default": "offline rules first; API only for ambiguity, mapping, OCR, or explicit reasoning",
        "reasoning_configured": "yes" if health.reasoning.configured else "no",
        "vision_configured": "yes" if health.vision.configured else "no",
        "reasoning_model": health.reasoning.model,
        "vision_model": health.vision.model,
    }
    if not input_path.is_dir():
        return ClientPackageDiagnostics(
            input_path=input_path,
            mode="missing_path",
            can_generate=False,
            confidence=0.0,
            missing_required=[f"Folder does not exist: {input_path}"],
            checks=[DiagnosticCheck("path_exists", False, str(input_path), 0.0)],
            provider_summary=provider_summary,
        )

    project = inspect_workpaper_project(input_path)
    cash = inspect_cash_materials(input_path)
    files = _scan_files(input_path)
    checks = _build_checks(project, cash, files)
    passed_weight = sum(check.confidence for check in checks if check.passed)
    total_weight = sum(max(check.confidence, 0.1) for check in checks) or 1.0
    confidence = min(1.0, passed_weight / total_weight)

    if project.is_ready:
        mode = "full_project"
        can_generate = True
        project_dir_resolved = project.project_dir
        cash_dir = project.cash_materials_dir
        missing = []
    elif cash.is_ready:
        mode = "single_cash"
        can_generate = True
        project_dir_resolved = None
        cash_dir = cash.materials_dir
        missing = []
    else:
        mode = "diagnostic_only"
        can_generate = False
        project_dir_resolved = project.project_dir
        cash_dir = project.cash_materials_dir
        missing = project.missing_required or cash.missing_required

    return ClientPackageDiagnostics(
        input_path=input_path,
        mode=mode,
        can_generate=can_generate,
        confidence=round(confidence, 3),
        project_dir=project_dir_resolved,
        cash_materials_dir=cash_dir,
        missing_required=missing,
        files=files,
        checks=checks,
        provider_summary=provider_summary,
        selected_cash_materials=_cash_summary(cash),
        cash_candidates=_cash_candidates(cash),
        agent_used=cash.agent_used,
        agent_reason=cash.agent_reason,
    )


def run_workpaper_stress_suite(
    fixtures_dir: str | Path,
    focus: Iterable[str] = ("C", "D", "EXP"),
) -> StressTestResult:
    root = _clean_path(fixtures_dir)
    focus_tuple = tuple(focus)
    cases: list[StressCaseResult] = []
    if not root.is_dir():
        return StressTestResult(
            fixtures_dir=root,
            focus=focus_tuple,
            case_count=0,
            runnable_count=0,
            success_count=0,
            invalid_handled_count=0,
            success_rate=0.0,
            cases=[],
        )

    case_dirs = [path for path in sorted(root.iterdir()) if path.is_dir() and not path.name.startswith("_")]
    output_base = Path("run_out") / f"stress_{safe_filename(root.name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    for case_dir in case_dirs:
        try:
            diag = diagnose_client_package(case_dir)
            if not diag.can_generate:
                cases.append(
                    StressCaseResult(
                        case_dir=case_dir,
                        mode=diag.mode,
                        can_generate=False,
                        success=True,
                        missing_required=diag.missing_required,
                        diagnostics_confidence=diag.confidence,
                    )
                )
                continue

            output_dir = output_base / safe_filename(case_dir.name)
            if diag.mode == "full_project":
                result = run_auto_workpaper_suite(diag.project_dir or case_dir, output_dir=output_dir)
                generated = list(result.workbooks)
                required = _required_codes_for_focus(focus_tuple)
                success = all(code in result.workbooks for code in required) and not any(
                    "Traceback" in error for error in result.errors
                )
                error = "; ".join(result.errors)
            else:
                cash_result = run_auto_cash_case(diag.cash_materials_dir or case_dir, case_dir=output_dir / "_cash_case")
                generated = ["C"] if cash_result.success else []
                success = "C" in generated and set(_required_codes_for_focus(focus_tuple)).issubset({"C"})
                error = cash_result.error
            cases.append(
                StressCaseResult(
                    case_dir=case_dir,
                    mode=diag.mode,
                    can_generate=True,
                    success=success,
                    generated_codes=generated,
                    diagnostics_confidence=diag.confidence,
                    output_dir=output_dir,
                    error=error,
                )
            )
        except Exception as exc:  # pragma: no cover - stress harness must keep going
            cases.append(
                StressCaseResult(
                    case_dir=case_dir,
                    mode="crashed",
                    can_generate=False,
                    success=False,
                    error=str(exc),
                )
            )

    runnable = [case for case in cases if case.can_generate]
    success_count = len([case for case in runnable if case.success])
    invalid_handled = len([case for case in cases if not case.can_generate and case.success])
    return StressTestResult(
        fixtures_dir=root,
        focus=focus_tuple,
        case_count=len(cases),
        runnable_count=len(runnable),
        success_count=success_count,
        invalid_handled_count=invalid_handled,
        success_rate=round(success_count / len(runnable), 3) if runnable else 0.0,
        cases=cases,
    )


def _build_checks(project, cash, files: list[DiagnosticFile]) -> list[DiagnosticCheck]:
    checks = [
        DiagnosticCheck("project_structure", project.is_ready, "; ".join(project.missing_required) or "full project detected", 1.0),
        DiagnosticCheck("cash_package", cash.is_ready, "; ".join(cash.missing_required) or "cash package detected", 0.8),
        DiagnosticCheck(
            "cash_material_coherence",
            cash.is_ready and not _looks_cross_project(cash),
            f"selected={cash.materials_dir}; confidence={cash.confidence:.0%}; agent_used={cash.agent_used}",
            0.9,
        ),
        _column_check("trial_balance_schema", files, "trial_balance", ("account_code", "account_name", "ending_balance")),
        _column_check("journal_schema", files, "journal", ("date", "account_code", "amount")),
        DiagnosticCheck(
            "provider_default_offline",
            True,
            "Default path uses pdf-text/offline extraction and deterministic Python rules.",
            0.5,
        ),
    ]
    return checks


def _cash_summary(cash) -> dict[str, str]:
    return {
        "materials_dir": str(cash.materials_dir),
        "trial_balance": str(cash.trial_balance or ""),
        "journal": str(cash.journal or ""),
        "bank_statement": str(cash.bank_statement or ""),
        "template": str(cash.template or ""),
        "confidence": f"{cash.confidence:.3f}",
    }


def _cash_candidates(cash) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in cash.candidate_sets[:8]:
        rows.append(
            {
                "root": str(item.root),
                "score": item.score,
                "confidence": item.confidence,
                "trial_balance": str(item.discovery.trial_balance or ""),
                "journal": str(item.discovery.journal or ""),
                "bank_statement": str(item.discovery.bank_statement or ""),
                "template": str(item.discovery.template or ""),
                "reasons": "; ".join(item.reasons),
            }
        )
    return rows


def _looks_cross_project(cash) -> bool:
    roots = []
    for path in [cash.trial_balance, cash.journal, cash.template]:
        if path:
            roots.append(_project_hint(Path(path)))
    return len(set(root for root in roots if root)) > 1


def _project_hint(path: Path) -> str:
    parts = path.parts
    for part in parts:
        if part.startswith("audit_sim_"):
            return part
    return str(path.parent)


def _column_check(
    name: str,
    files: list[DiagnosticFile],
    category: str,
    expected_columns: tuple[str, ...],
) -> DiagnosticCheck:
    candidates = [file for file in files if file.category == category and file.readable]
    for file in candidates:
        missing = [column for column in expected_columns if column not in file.matched_columns]
        if not missing:
            return DiagnosticCheck(name, True, f"{file.path.name}: {file.matched_columns}", 0.8)
    if candidates:
        best = max(candidates, key=lambda file: len(file.matched_columns))
        missing = [column for column in expected_columns if column not in best.matched_columns]
        return DiagnosticCheck(name, False, f"{best.path.name} missing aliases: {', '.join(missing)}", 0.8)
    return DiagnosticCheck(name, False, f"No readable {category} file detected", 0.8)


def _scan_files(root: Path) -> list[DiagnosticFile]:
    files: list[DiagnosticFile] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".csv", ".pdf"}:
            continue
        category = _classify_file(path)
        if path.suffix.lower() in {".xlsx", ".xlsm", ".csv"}:
            files.append(_inspect_structured_file(path, category))
        else:
            files.append(DiagnosticFile(path=path, category=category, readable=True))
    return files[:300]


def _inspect_structured_file(path: Path, category: str) -> DiagnosticFile:
    try:
        sheet_name, headers = _headers_for(path)
        return DiagnosticFile(
            path=path,
            category=category,
            readable=True,
            sheet_name=sheet_name,
            headers=headers,
            matched_columns=_match_headers(headers),
        )
    except Exception as exc:
        return DiagnosticFile(path=path, category=category, readable=False, error=str(exc))


def _headers_for(path: Path) -> tuple[str, list[str]]:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, encoding="utf-8-sig", nrows=0)
        return "", [str(column) for column in df.columns]

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        best_sheet = ""
        best_headers: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if len(values) > len(best_headers):
                    best_sheet = ws.title
                    best_headers = values
            if len(best_headers) >= 3:
                break
        return best_sheet, best_headers
    finally:
        wb.close()


def _match_headers(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers}
    matched: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = _normalize_header(alias)
            for norm_header, original in normalized.items():
                if norm_alias == norm_header or norm_alias in norm_header or norm_header in norm_alias:
                    matched[canonical] = original
                    break
            if canonical in matched:
                break
    return matched


def _classify_file(path: Path) -> str:
    text = f"{path.name} {path.parent.name}".lower()
    for category, keywords in FILE_CLASS_TERMS.items():
        if any(keyword.lower() in text for keyword in keywords):
            return category
    return "other"


def _required_codes_for_focus(focus: tuple[str, ...]) -> tuple[str, ...]:
    aliases = {"EXP": "EXP10", "D": "D10"}
    return tuple(aliases.get(code.upper(), code.upper()) for code in focus)


def _normalize_header(value: str) -> str:
    return str(value).strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _clean_path(value: str | Path) -> Path:
    text = str(value).strip().strip("\ufeff").strip()
    text = text.strip("`").strip().strip('"').strip("'").strip()
    if text.lower().startswith("file://"):
        text = text[7:]
    return Path(text).expanduser()
