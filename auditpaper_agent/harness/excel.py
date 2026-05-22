from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from auditpaper_agent.contracts import HarnessResult, ProvenanceEntry, WritePlan
from auditpaper_agent.trace import AuditTracer, ensure_tracer


def apply_write_plan(
    template_path: str | Path,
    output_path: str | Path,
    write_plan: WritePlan,
    tracer: AuditTracer | None = None,
) -> HarnessResult:
    """Copy template and apply a validated write plan."""
    tracer = ensure_tracer(tracer)
    template = Path(template_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    tracer.emit("Harness", "copied template workbook", f"{template} -> {output}")

    allowed = {sheet: {cell.upper() for cell in cells} for sheet, cells in write_plan.allowed_cells.items()}
    wb = load_workbook(output)
    provenance: list[ProvenanceEntry] = []
    try:
        for command in write_plan.commands:
            tracer.emit("Harness", "write requested", f"{command.sheet_name}!{command.cell} purpose={command.purpose}")
            if command.sheet_name not in wb.sheetnames:
                tracer.emit("Harness", "write rejected: sheet not found", command.sheet_name)
                raise ValueError(f"Sheet not found: {command.sheet_name}")
            if command.cell.upper() not in allowed.get(command.sheet_name, set()):
                tracer.emit("Harness", "write rejected: outside whitelist", f"{command.sheet_name}!{command.cell}")
                raise ValueError(f"Write outside whitelist: {command.sheet_name}!{command.cell}")
            tracer.emit("Harness", "whitelist check passed", f"{command.sheet_name}!{command.cell}")
            ws = wb[command.sheet_name]
            cell_addr = _merged_anchor(ws, command.cell.upper())
            cell = ws[cell_addr]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                tracer.emit("Harness", "write rejected: formula cell", f"{command.sheet_name}!{cell_addr}")
                raise ValueError(f"Refusing to overwrite formula cell: {command.sheet_name}!{cell_addr}")
            tracer.emit("Harness", "formula check passed", f"{command.sheet_name}!{cell_addr}")
            cell.value = _excel_value(command.value)
            cell.comment = Comment(
                text=(
                    f"AuditPaper-Agent\n"
                    f"Purpose: {command.purpose}\n"
                    f"Source: {command.source.label()}\n"
                    f"SHA256: {command.source.document_hash}"
                ),
                author="AuditPaper-Agent",
            )
            tracer.emit("Harness", "cell written with provenance comment", f"{command.sheet_name}!{cell_addr} source={command.source.document_id}")
            provenance.append(
                ProvenanceEntry(
                    sheet_name=command.sheet_name,
                    cell=cell_addr,
                    value_repr=repr(command.value),
                    purpose=command.purpose,
                    source=command.source,
                )
            )
        wb.save(output)
        tracer.emit("Harness", "workbook saved", str(output))
    finally:
        wb.close()
    return HarnessResult(output_path=str(output), commands_applied=len(provenance), provenance=provenance)


def _merged_anchor(ws: Worksheet, addr: str) -> str:
    for merged_range in ws.merged_cells.ranges:
        if addr in merged_range:
            return merged_range.start_cell.coordinate
    return addr


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    return value
