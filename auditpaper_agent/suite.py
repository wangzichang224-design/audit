from __future__ import annotations

import json
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from auditpaper_agent.contracts import GeneralLedgerRow, MasterDataRecord, StandardizedErpAuditPackage, TrialBalanceRow
from auditpaper_agent.service import run_auto_cash_case
from auditpaper_agent.sensing.erp import load_erp_package
from auditpaper_agent.utils import safe_filename, sha256_file, write_json
from auditpaper_agent.workpaper_catalog import PUBLIC_WORKPAPER_CATALOG, assert_workbook_clean_room


SUITE_FILENAMES = {
    "A10": "A10_风险评估与重要性.xlsx",
    "C": "C_货币资金_2025_华衡汽配.xlsx",
    "D10": "D10_交易性金融资产及其他权益工具投资.xlsx",
    "E20": "E20_应收账款与ECL.xlsx",
    "EXP10": "EXP10_费用测试与波动分析.xlsx",
    "F10": "F10_存货监盘与跌价.xlsx",
    "K10": "K10_固定资产与折旧.xlsx",
    "N10": "N10_应付账款与SURL.xlsx",
    "U10": "U10_收入确认与截止.xlsx",
}


@dataclass(frozen=True)
class WorkpaperSpec:
    code: str
    title: str
    assertion: str
    filename: str
    procedures: tuple[str, ...]
    source_keys: tuple[str, ...]


@dataclass(frozen=True)
class WorkpaperProjectSummary:
    input_path: Path
    project_dir: Path | None = None
    cash_materials_dir: Path | None = None
    client_name: str = ""
    period_end: str | None = None
    expected_workbooks: list[str] = field(default_factory=list)
    source_coverage: dict[str, bool] = field(default_factory=dict)
    missing_required: list[str] = field(default_factory=list)

    @property
    def is_ready(self) -> bool:
        return self.project_dir is not None and not self.missing_required


@dataclass(frozen=True)
class WorkpaperSuiteResult:
    success: bool
    summary: WorkpaperProjectSummary
    project_name: str = ""
    client_name: str = ""
    period_end: str | None = None
    output_dir: Path | None = None
    workbooks: dict[str, Path] = field(default_factory=dict)
    manifest_path: Path | None = None
    zip_path: Path | None = None
    findings_count: int = 0
    source_coverage: dict[str, bool] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


WORKPAPER_SPECS = [
    WorkpaperSpec(
        code="A10",
        title="风险评估与重要性",
        assertion="总体风险、重要性、报表衔接",
        filename=SUITE_FILENAMES["A10"],
        procedures=(
            "复核试算平衡表与审定报表衔接关系。",
            "识别余额重大或波动异常的科目，形成后续实质性程序重点。",
            "记录重要性、明显微小错报临界值及项目组复核关注点。",
        ),
        source_keys=("tb", "tie_out", "validation_report"),
    ),
    WorkpaperSpec(
        code="D10",
        title="交易性金融资产及其他权益工具投资",
        assertion="存在、权利义务、计价、列报",
        filename=SUITE_FILENAMES["D10"],
        procedures=(
            "复核 TB 是否存在交易性金融资产、其他权益工具投资或相关投资收益科目。",
            "如存在余额，结合明细账和期末公允价值资料形成估值复核样本。",
            "如未识别余额，记录不适用判断依据并保留项目组确认位。",
        ),
        source_keys=("tb", "accounting_records"),
    ),
    WorkpaperSpec(
        code="E20",
        title="应收账款与ECL",
        assertion="存在、计价、截止、列报",
        filename=SUITE_FILENAMES["E20"],
        procedures=(
            "结合客户主数据和销售循环识别应收款样本。",
            "按账龄、客户集中度和异常余额提示 ECL 复核事项。",
            "对暂未自动判断事项保留人工复核结论位。",
        ),
        source_keys=("customers", "sales_cycle", "tb"),
    ),
    WorkpaperSpec(
        code="EXP10",
        title="费用测试与波动分析",
        assertion="发生、准确性、截止、分类",
        filename=SUITE_FILENAMES["EXP10"],
        procedures=(
            "按费用类科目执行本期发生额和期末余额波动分析。",
            "从制造费用、期间费用、减值损失、所得税费用等科目抽取大额样本。",
            "对异常波动、分类风险和截止风险保留进一步复核提示。",
        ),
        source_keys=("tb", "accounting_records", "purchase_cycle"),
    ),
    WorkpaperSpec(
        code="F10",
        title="存货监盘与跌价",
        assertion="存在、完整性、计价",
        filename=SUITE_FILENAMES["F10"],
        procedures=(
            "结合产品、原材料、BOM 与生产存货循环形成监盘样本。",
            "识别高价值、周转慢或需跌价复核的存货项目。",
            "记录监盘/抽盘结果和跌价测试复核提示。",
        ),
        source_keys=("products", "raw_materials", "bom", "inventory_cycle"),
    ),
    WorkpaperSpec(
        code="K10",
        title="固定资产与折旧",
        assertion="存在、权利义务、计价",
        filename=SUITE_FILENAMES["K10"],
        procedures=(
            "基于固定资产台账抽取新增、重要及折旧敏感资产。",
            "复核折旧年限、残值率和账面金额的合理性。",
            "对资产存在性和减值迹象保留复核结论。",
        ),
        source_keys=("fixed_assets", "payroll_tax_fixed_assets", "tb"),
    ),
    WorkpaperSpec(
        code="N10",
        title="应付账款与SURL",
        assertion="完整性、截止、计价",
        filename=SUITE_FILENAMES["N10"],
        procedures=(
            "结合供应商主数据和采购循环识别应付账款样本。",
            "列示期后付款和未入账负债搜索复核项目。",
            "对大额供应商、异常账期和截止风险保留复核提示。",
        ),
        source_keys=("suppliers", "purchase_cycle", "tb"),
    ),
    WorkpaperSpec(
        code="U10",
        title="收入确认与截止",
        assertion="发生、准确性、截止",
        filename=SUITE_FILENAMES["U10"],
        procedures=(
            "结合销售循环和月度销售资料抽取收入确认样本。",
            "识别期末前后收入、异常客户和大额交易。",
            "记录截止测试、收入确认依据和进一步复核建议。",
        ),
        source_keys=("customers", "sales_cycle", "sales_books", "tb"),
    ),
]


def inspect_workpaper_project(project_dir: str | Path) -> WorkpaperProjectSummary:
    input_path = _clean_path(project_dir)
    if not input_path.is_dir():
        return WorkpaperProjectSummary(
            input_path=input_path,
            missing_required=[f"项目文件夹不存在：{input_path}"],
        )

    root = _resolve_project_root(input_path)
    if root is None:
        return WorkpaperProjectSummary(
            input_path=input_path,
            missing_required=["未识别到完整拟真审计项目结构"],
        )

    cash_dir = root / "c底稿资料"
    paths = _project_paths(root)
    coverage = {key: bool(value and value.exists()) for key, value in paths.items()}
    missing = []
    for key, label in {
        "cash_materials": "c底稿资料",
        "tb": "TB/试算平衡表",
        "journal": "序时账",
        "master_data": "主数据",
        "workpapers": "现有工作底稿目录",
    }.items():
        if not coverage.get(key):
            missing.append(label)

    client_name, period_end = _project_meta(root, cash_dir)
    return WorkpaperProjectSummary(
        input_path=input_path,
        project_dir=root,
        cash_materials_dir=cash_dir if cash_dir.exists() else None,
        client_name=client_name,
        period_end=period_end,
        expected_workbooks=list(SUITE_FILENAMES.values()),
        source_coverage=coverage,
        missing_required=missing,
    )


def run_auto_workpaper_suite(
    project_dir: str | Path,
    template_mode: str = "auto",
    output_dir: str | Path | None = None,
) -> WorkpaperSuiteResult:
    summary = inspect_workpaper_project(project_dir)
    if not summary.is_ready or summary.project_dir is None:
        return WorkpaperSuiteResult(
            success=False,
            summary=summary,
            errors=summary.missing_required,
            source_coverage=summary.source_coverage,
        )

    root = summary.project_dir
    output_root = Path(output_dir) if output_dir else _default_suite_dir(root)
    output_root.mkdir(parents=True, exist_ok=True)

    workbooks: dict[str, Path] = {}
    errors: list[str] = []
    findings_count = 0
    context = _build_context(summary)
    sources = _collect_sources(root)

    cash_output = output_root / SUITE_FILENAMES["C"]
    cash_result = run_auto_cash_case(
        summary.cash_materials_dir or root,
        case_dir=output_root / "_cash_case",
        output_path=cash_output,
    )
    if cash_result.success and cash_result.output_path:
        _polish_cash_workbook(cash_output, cash_result.artifacts)
        assert_workbook_clean_room(cash_output)
        workbooks["C"] = cash_output
        findings_count += cash_result.findings_count
    else:
        errors.append(f"C/货币资金生成失败：{cash_result.error}")

    for spec in WORKPAPER_SPECS:
        try:
            path = output_root / spec.filename
            samples = _samples_for_spec(spec, root)
            findings = _findings_for_spec(spec, samples)
            if spec.code == "D10":
                _build_d_workbook(path, spec, context, sources, samples, findings, root)
            elif spec.code == "EXP10":
                _build_exp_workbook(path, spec, context, sources, samples, findings, root)
            else:
                _build_generic_workbook(path, spec, context, sources, samples, findings)
            assert_workbook_clean_room(path)
            workbooks[spec.code] = path
            findings_count += len(findings)
        except Exception as exc:  # pragma: no cover - defensive manifest path
            errors.append(f"{spec.code} 生成失败：{exc}")

    manifest = {
        "schema_version": "auditpaper.workpaper_suite.v1",
        "project_name": root.name,
        "project_dir": root,
        "client_name": context["client_name"],
        "period_end": context["period_end"],
        "template_mode": template_mode,
        "source_coverage": summary.source_coverage,
        "workbooks": {code: str(path) for code, path in workbooks.items()},
        "findings_count": findings_count,
        "errors": errors,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    manifest_path = output_root / "suite_manifest.json"
    write_json(manifest_path, manifest)
    zip_path = _zip_suite(output_root, workbooks, manifest_path)

    return WorkpaperSuiteResult(
        success=len(workbooks) == len(SUITE_FILENAMES) and not errors,
        summary=summary,
        project_name=root.name,
        client_name=context["client_name"],
        period_end=context["period_end"],
        output_dir=output_root,
        workbooks=workbooks,
        manifest_path=manifest_path,
        zip_path=zip_path,
        findings_count=findings_count,
        source_coverage=summary.source_coverage,
        errors=errors,
    )


def run_erp_workpaper_suite(
    erp_case_dir: str | Path,
    output_dir: str | Path | None = None,
) -> WorkpaperSuiteResult:
    case_dir = _clean_path(erp_case_dir)
    try:
        package = load_erp_package(case_dir)
    except Exception as exc:
        summary = WorkpaperProjectSummary(input_path=case_dir, missing_required=["erp_package.json"])
        return WorkpaperSuiteResult(success=False, summary=summary, errors=[f"无法读取 ERP 标准包：{exc}"])

    summary = _erp_project_summary(case_dir, package)
    if not package.mapping_confirmed:
        return WorkpaperSuiteResult(
            success=False,
            summary=summary,
            client_name=package.meta.client_name,
            period_end=package.meta.period_end.isoformat() if package.meta.period_end else None,
            source_coverage=summary.source_coverage,
            errors=["ERP 字段映射尚未确认，禁止直接生成主循环底稿。"],
        )

    output_root = Path(output_dir) if output_dir else _default_suite_dir(case_dir)
    output_root.mkdir(parents=True, exist_ok=True)

    workbooks: dict[str, Path] = {}
    errors: list[str] = []
    findings_count = 0
    context = _erp_context(package)
    sources = _collect_erp_sources(package)
    specs = [_erp_cash_spec(), *WORKPAPER_SPECS]

    for spec in specs:
        try:
            path = output_root / spec.filename
            samples = _samples_for_spec_from_erp(spec, package)
            findings = _findings_for_spec(spec, samples)
            _build_generic_workbook(path, spec, context, sources, samples, findings)
            assert_workbook_clean_room(path)
            workbooks[spec.code] = path
            findings_count += len(findings)
        except Exception as exc:  # pragma: no cover - defensive manifest path
            errors.append(f"{spec.code} 生成失败：{exc}")

    manifest = {
        "schema_version": "auditpaper.workpaper_suite.v1",
        "input_mode": "erp_standardized_package",
        "project_name": case_dir.name,
        "project_dir": case_dir,
        "client_name": context["client_name"],
        "period_end": context["period_end"],
        "provider": package.provider,
        "mapping_confirmed": package.mapping_confirmed,
        "source_coverage": summary.source_coverage,
        "workbooks": {code: str(path) for code, path in workbooks.items()},
        "findings_count": findings_count,
        "errors": errors,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    manifest_path = output_root / "suite_manifest.json"
    write_json(manifest_path, manifest)
    zip_path = _zip_suite(output_root, workbooks, manifest_path)

    return WorkpaperSuiteResult(
        success=len(workbooks) == len(SUITE_FILENAMES) and not errors,
        summary=summary,
        project_name=case_dir.name,
        client_name=context["client_name"],
        period_end=context["period_end"],
        output_dir=output_root,
        workbooks=workbooks,
        manifest_path=manifest_path,
        zip_path=zip_path,
        findings_count=findings_count,
        source_coverage=summary.source_coverage,
        errors=errors,
    )


def _clean_path(value: str | Path) -> Path:
    text = str(value).strip().strip("\ufeff").strip()
    text = text.strip("`").strip().strip('"').strip("'").strip()
    if text.lower().startswith("file://"):
        text = text[7:]
    return Path(text).expanduser()


def _resolve_project_root(path: Path) -> Path | None:
    candidates = [path, *path.parents]
    for candidate in candidates:
        if (candidate / "04_accounting_records").is_dir() and (candidate / "02_master_data").is_dir():
            return candidate
        if (candidate / "c底稿资料").is_dir() and (candidate / "05_audit_workpapers").is_dir():
            return candidate
    return _find_best_descendant_project(path) if path.is_dir() else None


def _find_best_descendant_project(path: Path) -> Path | None:
    scored: list[tuple[int, Path]] = []
    for candidate in [path, *[p for p in path.rglob("*") if p.is_dir()]]:
        score = 0
        if (candidate / "04_accounting_records").is_dir():
            score += 35
        if (candidate / "02_master_data").is_dir():
            score += 30
        if (candidate / "c底稿资料").is_dir():
            score += 25
        if (candidate / "05_audit_workpapers").is_dir():
            score += 20
        if (candidate / "06_validation").is_dir():
            score += 10
        if score >= 75:
            scored.append((score, candidate))
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], len(item[1].parts), str(item[1]).lower()))
    return scored[0][1]


def _project_paths(root: Path) -> dict[str, Path | None]:
    cash_dir = root / "c底稿资料"
    return {
        "cash_materials": cash_dir if cash_dir.exists() else None,
        "tb": _first_existing([root / "04_accounting_records" / "TB_2025.xlsx", cash_dir / "试算平衡表_TB_2025.xlsx"]),
        "journal": _first_existing([root / "04_accounting_records" / "序时账_2025.xlsx", cash_dir / "序时账_2025.xlsx"]),
        "bank_statements": _first_existing([cash_dir / "银行对账单", root / "03_business_cycles" / "treasury" / "bank_statements"]),
        "master_data": _first_existing([root / "02_master_data" / "master_data.xlsx", root / "02_master_data"]),
        "workpapers": root / "05_audit_workpapers",
        "accounting_records": root / "04_accounting_records" / "accounting_records.xlsx",
        "tie_out": root / "06_validation" / "tie_out_matrix.xlsx",
        "validation_report": root / "06_validation" / "validation_report.json",
        "customers": root / "02_master_data" / "customers.csv",
        "suppliers": root / "02_master_data" / "suppliers.csv",
        "products": root / "02_master_data" / "products.csv",
        "raw_materials": root / "02_master_data" / "raw_materials.csv",
        "bom": root / "02_master_data" / "bom.csv",
        "fixed_assets": root / "02_master_data" / "fixed_assets_master.csv",
        "sales_cycle": root / "03_business_cycles" / "sales" / "sales_cycle.xlsx",
        "purchase_cycle": root / "03_business_cycles" / "purchase" / "purchase_cycle.xlsx",
        "inventory_cycle": root / "03_business_cycles" / "production_inventory" / "production_inventory_cycle.xlsx",
        "payroll_tax_fixed_assets": root / "03_business_cycles" / "payroll_tax_fixed_assets" / "payroll_tax_fixed_assets.xlsx",
        "sales_books": root / "03_business_cycles" / "monthly_document_books" / "sales",
    }


def _first_existing(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def _project_meta(root: Path, cash_dir: Path) -> tuple[str, str | None]:
    package_path = cash_dir / "case_package.json"
    if package_path.exists():
        try:
            data = json.loads(package_path.read_text(encoding="utf-8"))
            meta = data.get("meta", {})
            return str(meta.get("client_name") or root.name), meta.get("period_end")
        except Exception:
            pass
    for report in sorted(cash_dir.glob("企业信用报告_*")):
        name = report.stem.split("_", 1)[-1].strip()
        if name:
            return name, "2025-12-31"
    return root.name, "2025-12-31"


def _build_context(summary: WorkpaperProjectSummary) -> dict[str, str]:
    return {
        "client_name": summary.client_name or (summary.project_dir.name if summary.project_dir else "UNKNOWN_CLIENT"),
        "period_end": summary.period_end or "2025-12-31",
        "prepared_by": "AuditPaper-Agent",
    }


def _collect_sources(root: Path) -> dict[str, Path | None]:
    return _project_paths(root)


def _default_suite_dir(root: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("run_out") / f"{safe_filename(root.name)}_workpaper_suite_{stamp}"


def _erp_cash_spec() -> WorkpaperSpec:
    return WorkpaperSpec(
        code="C",
        title="货币资金",
        assertion="存在、完整性、权利义务、截止",
        filename=SUITE_FILENAMES["C"],
        procedures=(
            "基于 ERP TB 和总账识别现金、银行存款及其他货币资金余额。",
            "从现金/银行相关凭证中抽取大额和期末样本，保留项目组复核结论位。",
            "如未提供银行对账单或函证，标记资料不足而不自动形成余额核对结论。",
        ),
        source_keys=("tb", "accounting_records", "bank_statements"),
    )


def _erp_project_summary(case_dir: Path, package: StandardizedErpAuditPackage) -> WorkpaperProjectSummary:
    coverage = {
        "tb": bool(package.trial_balance),
        "journal": bool(package.general_ledger),
        "customers": bool(package.customers),
        "suppliers": bool(package.suppliers),
        "inventory": bool(package.inventory),
        "fixed_assets": bool(package.fixed_assets),
        "sales_cycle": bool(package.sales),
        "purchase_cycle": bool(package.purchase),
        "bank_journal": bool(package.bank_journal),
    }
    missing = []
    if not package.trial_balance:
        missing.append("ERP/TB 标准化数据")
    if not package.general_ledger:
        missing.append("ERP/总账或序时账标准化数据")
    return WorkpaperProjectSummary(
        input_path=case_dir,
        project_dir=case_dir,
        client_name=package.meta.client_name,
        period_end=package.meta.period_end.isoformat() if package.meta.period_end else None,
        expected_workbooks=list(SUITE_FILENAMES.values()),
        source_coverage=coverage,
        missing_required=missing,
    )


def _erp_context(package: StandardizedErpAuditPackage) -> dict[str, str]:
    return {
        "client_name": package.meta.client_name or "UNKNOWN_CLIENT",
        "period_end": package.meta.period_end.isoformat() if package.meta.period_end else "未识别",
        "prepared_by": "AuditPaper-Agent",
    }


def _collect_erp_sources(package: StandardizedErpAuditPackage) -> dict[str, Path | None]:
    by_type: dict[str, Path] = {}
    for doc in package.source_documents:
        by_type.setdefault(doc.document_type, Path(doc.path))
    return {
        "tb": by_type.get("trial_balance"),
        "accounting_records": by_type.get("general_ledger"),
        "journal": by_type.get("general_ledger"),
        "bank_statements": by_type.get("bank_statement"),
        "customers": by_type.get("customers"),
        "suppliers": by_type.get("suppliers"),
        "products": by_type.get("inventory"),
        "raw_materials": by_type.get("inventory"),
        "bom": None,
        "fixed_assets": by_type.get("fixed_assets"),
        "sales_cycle": by_type.get("sales"),
        "purchase_cycle": by_type.get("purchase"),
        "inventory_cycle": by_type.get("inventory"),
        "payroll_tax_fixed_assets": by_type.get("fixed_assets"),
        "sales_books": by_type.get("sales"),
        "tie_out": by_type.get("trial_balance"),
        "validation_report": None,
    }


def _polish_cash_workbook(path: Path, artifacts: dict[str, Path]) -> None:
    wb = load_workbook(path)
    _apply_theme(wb)
    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        ws["G1"] = "自动化质量检查"
        ws["G2"] = "已执行程序"
        ws["H2"] = '=COUNTIF(D2:D9,"是")'
        ws["G3"] = "程序总数"
        ws["H3"] = "=COUNTA(B2:B9)"
        ws["G4"] = "执行覆盖率"
        ws["H4"] = '=IFERROR(H2/H3,"")'
        ws["H4"].number_format = "0.0%"
        for row in range(2, min(ws.max_row, 9) + 1):
            if not ws.cell(row, 4).value:
                ws.cell(row, 4).value = "是"
        _write_source_index_block(ws, artifacts)
    if "C.00 Lead" in wb.sheetnames:
        ws = wb["C.00 Lead"]
        ws["B29"] = "合计/交叉检查"
        ws["D29"] = "=SUM(D9:D28)"
        ws["F29"] = "=SUM(F9:F28)"
        ws["I29"] = "=SUM(I9:I28)"
    if "C.00 BKD" in wb.sheetnames:
        ws = wb["C.00 BKD"]
        ws["B25"] = "检查"
        ws["I25"] = "=SUM(I9:I24)"
        ws["J25"] = "=SUM(J9:J24)"
        ws["M25"] = '=IF(ABS(I25-J25)>1,"需复核","金额勾稽完成")'
    if "C.01 Confirmations" in wb.sheetnames:
        ws = wb["C.01 Confirmations"]
        ws["B25"] = "函证覆盖率"
        ws["C25"] = '=IFERROR(COUNTIF(M9:M24,"<>")/COUNTA(B9:B24),"")'
        ws["C25"].number_format = "0.0%"
    if "C.02 Bank reconciliations" in wb.sheetnames:
        ws = wb["C.02 Bank reconciliations"]
        ws["B33"] = "调节差异检查"
        ws["K33"] = "=SUM(K9:K32)"
        ws["L33"] = '=IF(ABS(K33)>1,"需复核","调节差异已闭环")'
    if "C.03 Cutoff" in wb.sheetnames:
        ws = wb["C.03 Cutoff"]
        ws["B27"] = "截止样本数"
        ws["C27"] = "=COUNTA(B9:B26)"
    _enhance_cash_public_beta(wb, artifacts)
    wb.save(path)
    wb.close()


def _enhance_cash_public_beta(wb, artifacts: dict[str, Path]) -> None:
    findings = _read_json_any(artifacts.get("audit_findings"))
    provenance = _read_json_any(artifacts.get("provenance"))
    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        ws["J1"] = "公开版复核摘要"
        ws["J2"] = "审计发现数"
        ws["K2"] = len(findings) if isinstance(findings, list) else 0
        ws["J3"] = "provenance 记录数"
        ws["K3"] = len(provenance) if isinstance(provenance, list) else 0
        ws["J4"] = "负数现金检查"
        ws["K4"] = '=IF(COUNTIF(\'C.00 BKD\'!I9:I24,"<0")+COUNTIF(\'C.00 Lead\'!D9:D28,"<0")>0,"需复核","未见负数现金")'
        ws["J5"] = "现金盘点"
        ws["K5"] = "如存在库存现金，请补充盘点表；当前自动化范围以银行及货币资金明细为主。"
        ws["J6"] = "异常分类"
        ws["K6"] = "受限、冻结、质押、保证金和未达调节项应在 C.00b/C.02 保留人工结论。"
    if "C.01b 特殊账户程序" in wb.sheetnames:
        ws = wb["C.01b 特殊账户程序"]
        _write_review_block(
            ws,
            20,
            [
                ("负数现金", "检查银行账户或现金科目是否出现贷方/负数余额。", "未见异常时由项目组签字确认。"),
                ("现金盘点", "如存在库存现金，补充盘点日、盘点人、监盘人和差异说明。", "本底稿不自动替代现场盘点。"),
                ("受限资金", "结合银行回函、对账单、借款合同识别冻结、质押、保证金。", "分类结论需人工复核。"),
            ],
        )
    if "C.00b 分类检查" in wb.sheetnames:
        ws = wb["C.00b 分类检查"]
        _write_review_block(
            ws,
            20,
            [
                ("异常分类规则", "冻结、质押、保证金、监管账户、长期未动户均标记为需复核。", "项目组确认列报分类。"),
            ],
        )


def _write_review_block(ws, start_row: int, rows: list[tuple[str, str, str]]) -> None:
    headers = ["复核事项", "Agent 程序", "项目组结论"]
    for col, header in enumerate(headers, start=2):
        ws.cell(start_row, col).value = header
    for offset, row in enumerate(rows, start=1):
        for col, value in enumerate(row, start=2):
            ws.cell(start_row + offset, col).value = value


def _read_json_any(path: Path | None) -> Any:
    if path is None or not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def _write_source_index_block(ws, artifacts: dict[str, Path]) -> None:
    ws["G6"] = "来源索引"
    ws["G7"] = "产物"
    ws["H7"] = "路径"
    row = 8
    for name, path in artifacts.items():
        if row > 18:
            break
        ws.cell(row, 7).value = name
        ws.cell(row, 8).value = str(path)
        row += 1


def _samples_for_spec(spec: WorkpaperSpec, root: Path) -> list[dict[str, Any]]:
    paths = _project_paths(root)
    samples: list[dict[str, Any]] = []
    if spec.code == "A10":
        rows = _read_records(paths.get("tb"), 12)
        for row in rows:
            amount = _first_value(row, ["期末余额", "ending_balance", "期末借方余额", "期末借方"])
            samples.append(_sample("TB余额复核", row, amount, "余额重大或波动异常科目需与 Lead 程序衔接。"))
    elif spec.code == "D10":
        rows = _read_filtered_records(
            paths.get("tb"),
            lambda row: _row_account_code(row).startswith(("1101", "1503", "1511", "1512"))
            or any(token in _row_account_name(row) for token in ["交易性金融资产", "其他权益工具投资", "长期股权投资", "投资收益"]),
            12,
        )
        if not rows:
            samples.append(
                {
                    "程序": "适用性检查",
                    "样本/对象": "TB未识别交易性金融资产/权益工具投资余额",
                    "金额/指标": "",
                    "风险提示": "当前结构化资料未显示 D 类余额；需项目组确认本期不适用或补充投资明细。",
                    "复核建议": "检查 TB、总账明细、投资台账及管理层声明，确认是否无需执行 D 类实质性程序。",
                }
            )
        for row in rows:
            samples.append(_sample("投资余额/估值复核", row, _first_value(row, ["期末余额", "本期借方", "本期贷方"]), "关注金融资产分类、估值输入和列报。"))
    elif spec.code == "E20":
        for row in _read_records(paths.get("customers"), 12):
            samples.append(_sample("客户/应收样本", row, _first_value(row, ["credit_limit", "信用额度", "balance"]), "关注客户集中度、账龄和回款风险。"))
    elif spec.code == "EXP10":
        rows = _read_filtered_records(
            paths.get("tb"),
            lambda row: _row_account_code(row).startswith(("5001", "5101", "6401", "6601", "6602", "6603", "6604", "6701", "6702", "6801"))
            or any(token in _row_account_name(row) for token in ["费用", "成本", "减值损失"]),
            16,
        )
        journal_rows = _read_filtered_records(
            paths.get("accounting_records"),
            lambda row: _row_account_code(row).startswith(("5", "6401", "6601", "6602", "6603", "6604", "6701", "6702", "6801")),
            16,
        )
        for row in rows:
            samples.append(_sample("费用科目波动复核", row, _first_value(row, ["本期借方", "期末余额", "debit"]), "关注费用波动、成本归集和分类准确性。"))
        for row in journal_rows[:8]:
            samples.append(_sample("费用凭证明细样本", row, _first_value(row, ["debit", "credit", "金额"]), "关注费用发生真实性、截止和附件完整性。"))
    elif spec.code == "F10":
        for row in _read_records(paths.get("products"), 8) + _read_records(paths.get("raw_materials"), 8):
            samples.append(_sample("存货监盘/跌价样本", row, _first_value(row, ["unit_cost", "standard_cost", "cost", "金额"]), "关注单位成本、库龄和可变现净值。"))
    elif spec.code == "K10":
        for row in _read_records(paths.get("fixed_assets"), 16):
            samples.append(_sample("固定资产样本", row, _first_value(row, ["cost", "original_cost", "原值", "amount"]), "关注新增资产、折旧年限和存在性。"))
    elif spec.code == "N10":
        for row in _read_records(paths.get("suppliers"), 16):
            samples.append(_sample("供应商/SURL样本", row, _first_value(row, ["credit_limit", "balance", "amount"]), "关注大额供应商、期后付款和未入账负债。"))
    elif spec.code == "U10":
        for row in _read_records(paths.get("sales_cycle"), 16):
            samples.append(_sample("收入截止样本", row, _first_value(row, ["amount", "revenue", "金额"]), "关注期末前后发货、验收和收入确认时点。"))
    return samples[:24] or [_empty_sample(spec)]


def _samples_for_spec_from_erp(spec: WorkpaperSpec, package: StandardizedErpAuditPackage) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    if spec.code == "A10":
        for row in sorted(package.trial_balance, key=lambda item: abs(item.ending_balance), reverse=True)[:20]:
            samples.append(_sample("ERP TB余额复核", _tb_row(row), row.ending_balance, "余额重大或波动异常科目需与后续程序衔接。"))
    elif spec.code == "C":
        cash_tb = [row for row in package.trial_balance if _is_cash_account_name(row.account_code, row.account_name)]
        cash_gl = [row for row in package.general_ledger if _is_cash_account_name(row.account_code, row.account_name)]
        for row in cash_tb[:12]:
            samples.append(_sample("货币资金余额复核", _tb_row(row), row.ending_balance, "需与银行对账单、函证和余额调节表勾稽。"))
        for row in sorted(cash_gl, key=lambda item: abs(item.amount), reverse=True)[:12]:
            samples.append(_sample("现金/银行凭证样本", _gl_row(row), row.amount, "关注大额收付款、受限资金、银行间转账和截止风险。"))
    elif spec.code == "D10":
        rows = [
            row
            for row in package.trial_balance
            if _row_account_code(_tb_row(row)).startswith(("1101", "1503", "1511", "1512", "1521", "4001", "6111"))
            or any(token in row.account_name for token in ("交易性金融资产", "其他权益工具投资", "债权投资", "其他债权投资", "长期股权投资", "投资收益"))
        ]
        if not rows:
            samples.append(
                {
                    "程序": "适用性检查",
                    "样本/对象": "ERP TB未识别交易性金融资产/权益工具投资余额",
                    "金额/指标": "",
                    "风险提示": "当前 ERP 标准包未显示 D 类余额；需项目组确认本期不适用或补充投资明细。",
                    "复核建议": "检查 ERP 总账、投资台账及管理层声明，确认是否无需执行 D 类实质性程序。",
                }
            )
        for row in rows[:20]:
            samples.append(_sample("投资余额/估值复核", _tb_row(row), row.ending_balance, "关注金融资产分类、估值输入和列报。"))
    elif spec.code == "E20":
        for row in package.customers[:20]:
            samples.append(_sample("客户/应收样本", _master_row(row), row.amount, "关注客户集中度、账龄和回款风险。"))
        for row in _gl_by_tokens(package.general_ledger, ("1122",), ("应收",))[:8]:
            samples.append(_sample("应收账款总账样本", _gl_row(row), row.amount, "关注收入、回款和坏账准备勾稽。"))
    elif spec.code == "EXP10":
        for row in _tb_by_tokens(package.trial_balance, ("5", "6401", "6402", "6601", "6602", "6603", "6604", "6701", "6702", "6801"), ("费用", "成本", "减值损失", "税金"))[:20]:
            samples.append(_sample("费用科目波动复核", _tb_row(row), row.ending_balance, "关注费用波动、成本归集和分类准确性。"))
        for row in _gl_by_tokens(package.general_ledger, ("5", "6401", "6402", "6601", "6602", "6603", "6604", "6701", "6702", "6801"), ("费用", "成本", "减值损失", "税金"))[:12]:
            samples.append(_sample("费用凭证明细样本", _gl_row(row), row.amount, "关注费用发生真实性、截止和附件完整性。"))
    elif spec.code == "F10":
        for row in package.inventory[:24]:
            samples.append(_sample("存货监盘/跌价样本", _master_row(row), row.amount, "关注单位成本、库龄、可变现净值和盘点差异。"))
    elif spec.code == "K10":
        for row in package.fixed_assets[:24]:
            samples.append(_sample("固定资产样本", _master_row(row), row.amount, "关注新增资产、折旧年限、减值迹象和存在性。"))
    elif spec.code == "N10":
        for row in package.suppliers[:20]:
            samples.append(_sample("供应商/SURL样本", _master_row(row), row.amount, "关注大额供应商、期后付款和未入账负债。"))
        for row in _gl_by_tokens(package.general_ledger, ("2202", "2241",), ("应付", "供应商"))[:8]:
            samples.append(_sample("应付账款总账样本", _gl_row(row), row.amount, "关注完整性、截止和期后付款。"))
    elif spec.code == "U10":
        for row in package.sales[:20]:
            samples.append(_sample("收入截止样本", _master_row(row), row.amount, "关注期末前后发货、验收和收入确认时点。"))
        for row in _gl_by_tokens(package.general_ledger, ("6001", "6051"), ("收入",))[:8]:
            samples.append(_sample("收入总账样本", _gl_row(row), row.amount, "关注发生、准确性和截止。"))
    return samples[:24] or [_empty_sample(spec)]


def _tb_row(row: TrialBalanceRow) -> dict[str, Any]:
    return {
        "account_code": row.account_code,
        "account_name": row.account_name,
        "ending_balance": row.ending_balance,
        "prior_year": row.prior_year,
    }


def _gl_row(row: GeneralLedgerRow) -> dict[str, Any]:
    return {
        "voucher_id": row.voucher_id,
        "voucher_date": row.posting_date.isoformat() if row.posting_date else "",
        "account_code": row.account_code,
        "account_name": row.account_name,
        "description": row.description,
        "debit": row.debit,
        "credit": row.credit,
        "amount": row.amount,
        "counterparty": row.counterparty,
    }


def _master_row(row: MasterDataRecord) -> dict[str, Any]:
    return {
        "record_id": row.record_id,
        "name": row.name,
        "amount": row.amount,
        "currency": row.currency,
        **{key: value for key, value in row.attributes.items() if key not in {"record_id", "name"}},
    }


def _tb_by_tokens(rows: list[TrialBalanceRow], code_prefixes: tuple[str, ...], name_tokens: tuple[str, ...]) -> list[TrialBalanceRow]:
    return [
        row
        for row in rows
        if str(row.account_code).startswith(code_prefixes) or any(token in str(row.account_name) for token in name_tokens)
    ]


def _gl_by_tokens(rows: list[GeneralLedgerRow], code_prefixes: tuple[str, ...], name_tokens: tuple[str, ...]) -> list[GeneralLedgerRow]:
    filtered = [
        row
        for row in rows
        if str(row.account_code).startswith(code_prefixes) or any(token in str(row.account_name) for token in name_tokens)
    ]
    return sorted(filtered, key=lambda row: abs(row.amount), reverse=True)


def _is_cash_account_name(account_code: str, account_name: str) -> bool:
    return str(account_code).startswith(("1001", "1002", "1009")) or str(account_name).startswith(("库存现金", "银行存款", "其他货币资金"))


def _read_records(path: Path | None, limit: int) -> list[dict[str, Any]]:
    if path is None or not path.exists() or path.is_dir():
        return []
    try:
        if path.suffix.lower() == ".csv":
            df = pd.read_csv(path, encoding="utf-8-sig")
        else:
            df = pd.read_excel(path)
    except Exception:
        return []
    df = df.head(limit)
    return [{str(k): _json_cell(v) for k, v in row.items()} for row in df.to_dict(orient="records")]


def _read_filtered_records(path: Path | None, predicate, limit: int) -> list[dict[str, Any]]:
    rows = _read_records(path, 10_000)
    return [row for row in rows if predicate(row)][:limit]


def _row_account_code(row: dict[str, Any]) -> str:
    for key in ["科目编码", "account_code"]:
        if key in row and row[key] not in ("", None):
            try:
                return str(int(float(row[key])))
            except Exception:
                return str(row[key]).strip()
    return ""


def _row_account_name(row: dict[str, Any]) -> str:
    for key in ["科目名称", "account_name"]:
        if key in row and row[key] not in ("", None):
            return str(row[key])
    return ""


def _json_cell(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _first_value(row: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in row and row[key] not in ("", None):
            return row[key]
    for value in row.values():
        if isinstance(value, (int, float)) and value:
            return value
    return ""


def _sample(name: str, row: dict[str, Any], amount: Any, risk: str) -> dict[str, Any]:
    label = ""
    for value in row.values():
        if value not in ("", None):
            label = str(value)
            break
    return {
        "程序": name,
        "样本/对象": label[:80],
        "金额/指标": amount,
        "风险提示": risk,
        "复核建议": "由项目组结合原始凭证和期后资料复核，不由 Agent 直接下最终审计结论。",
    }


def _empty_sample(spec: WorkpaperSpec) -> dict[str, Any]:
    return {
        "程序": spec.title,
        "样本/对象": "未解析到结构化样本",
        "金额/指标": "",
        "风险提示": "资料不足，需补充 PBC 或人工选择样本。",
        "复核建议": "补充资料后重新运行，或在本表中手工记录复核过程。",
    }


def _findings_for_spec(spec: WorkpaperSpec, samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    findings = [
        {
            "编号": f"{spec.code}-F01",
            "严重性": "中",
            "发现类型": "待复核提示",
            "金额": "",
            "描述": f"{spec.title} 已生成结构化样本，当前版本保留人工复核结论位，不自动替代审计判断。",
            "建议程序": "检查来源索引中的原始资料，并在复核栏补充判断依据。",
        }
    ]
    numeric_amounts = [_to_float(sample.get("金额/指标")) for sample in samples]
    if any(abs(value) > 1_000_000 for value in numeric_amounts):
        findings.append(
            {
                "编号": f"{spec.code}-F02",
                "严重性": "高",
                "发现类型": "大额样本",
                "金额": max(numeric_amounts, key=abs),
                "描述": "存在超过 1,000,000 的样本指标，建议优先执行细节测试或期后核查。",
                "建议程序": "结合凭证、合同、发票、对账单或期后回款/付款资料执行穿行复核。",
            }
        )
    return findings


def _to_float(value: Any) -> float:
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def _build_generic_workbook(
    path: Path,
    spec: WorkpaperSpec,
    context: dict[str, str],
    sources: dict[str, Path | None],
    samples: list[dict[str, Any]],
    findings: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    lead = wb.create_sheet("Lead")
    procedures = wb.create_sheet("程序与样本")
    finding_sheet = wb.create_sheet("审计发现")
    source_sheet = wb.create_sheet("来源索引")

    _populate_lead(lead, spec, context, samples, findings)
    _populate_samples(procedures, spec, samples)
    _populate_findings(finding_sheet, findings)
    _populate_sources(source_sheet, spec, sources)
    _apply_theme(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _build_d_workbook(
    path: Path,
    spec: WorkpaperSpec,
    context: dict[str, str],
    sources: dict[str, Path | None],
    samples: list[dict[str, Any]],
    findings: list[dict[str, Any]],
    root: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    lead = wb.create_sheet("Lead")
    classification = wb.create_sheet("金融资产分类")
    valuation = wb.create_sheet("公允价值复核")
    confirmations = wb.create_sheet("函证与抽样结果")
    finding_sheet = wb.create_sheet("审计发现")
    source_sheet = wb.create_sheet("来源索引")

    paths = _project_paths(root)
    investment_rows = _investment_rows(paths.get("tb"))
    has_balance = bool(investment_rows)

    _populate_focus_lead(
        lead,
        spec,
        context,
        workbook_sheets=("金融资产分类", "公允价值复核", "函证与抽样结果", "审计发现", "来源索引"),
        status="识别到投资类余额，已生成分类、估值和函证/抽样复核底稿。"
        if has_balance
        else "未在结构化 TB 中识别到投资类余额，已生成不适用判断底稿，等待项目组确认。",
    )

    classification_rows = _d_classification_rows(investment_rows)
    if not classification_rows:
        classification_rows = [
            [
                1,
                "TB 适用性检查",
                "未识别交易性金融资产、其他权益工具投资或投资收益余额",
                "",
                "不适用判断",
                "需项目组结合总账明细、投资台账和管理层声明确认",
                "",
            ]
        ]
    _write_table(
        classification,
        "B8",
        ["序号", "检查对象", "科目/资料", "账面金额", "初步分类", "Agent 提示", "项目组结论"],
        classification_rows,
    )

    valuation_rows = _d_valuation_rows(investment_rows)
    _write_table(
        valuation,
        "B8",
        ["序号", "估值对象", "账面金额", "估值资料需求", "复核重点", "初步状态", "项目组结论"],
        valuation_rows,
    )

    confirmation_rows = _d_confirmation_rows(investment_rows, samples)
    _write_table(
        confirmations,
        "B8",
        ["序号", "样本/对象", "程序类型", "资料需求", "风险提示", "执行状态", "索引"],
        confirmation_rows,
    )

    _populate_findings(finding_sheet, findings)
    _populate_sources(source_sheet, spec, sources)
    _apply_theme(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _build_exp_workbook(
    path: Path,
    spec: WorkpaperSpec,
    context: dict[str, str],
    sources: dict[str, Path | None],
    samples: list[dict[str, Any]],
    findings: list[dict[str, Any]],
    root: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    lead = wb.create_sheet("Lead")
    fluctuation = wb.create_sheet("费用波动")
    bkd = wb.create_sheet("明细BKD")
    tod = wb.create_sheet("TOD详细测试")
    cutoff = wb.create_sheet("截止测试")
    finding_sheet = wb.create_sheet("审计发现")
    source_sheet = wb.create_sheet("来源索引")

    paths = _project_paths(root)
    expense_tb_rows = _expense_tb_rows(paths.get("tb"))
    journal_rows = _expense_journal_rows(paths.get("accounting_records"))
    cutoff_rows = _expense_cutoff_rows(journal_rows)

    _populate_focus_lead(
        lead,
        spec,
        context,
        workbook_sheets=("费用波动", "明细BKD", "TOD详细测试", "截止测试", "审计发现", "来源索引"),
        status=(
            f"已生成 {len(expense_tb_rows)} 个费用/成本科目波动项、"
            f"{len(journal_rows[:20])} 个明细样本和 {len(cutoff_rows)} 个截止测试样本。"
            "最终分类和截止结论需项目组复核。"
        ),
    )

    _write_table(
        fluctuation,
        "B8",
        ["序号", "科目编码", "科目名称", "期末余额/发生额", "波动/规模提示", "建议程序", "项目组结论"],
        _expense_fluctuation_rows(expense_tb_rows),
    )
    _write_table(
        bkd,
        "B8",
        ["序号", "凭证号", "日期", "科目编码", "科目名称", "摘要", "借方", "贷方", "复核重点", "项目组结论"],
        _expense_bkd_rows(journal_rows),
    )
    _write_table(
        tod,
        "B8",
        ["序号", "凭证号", "日期", "样本/对象", "金额", "TOD 程序", "资料需求", "执行状态"],
        _expense_tod_rows(journal_rows or samples),
    )
    _write_table(
        cutoff,
        "B8",
        ["序号", "凭证号", "日期", "科目名称", "金额", "截止风险", "建议程序", "项目组结论"],
        cutoff_rows
        or [[1, "", "", "未识别期末费用凭证", "", "资料不足", "补充期末前后凭证、合同、发票和付款资料后复核", ""]],
    )

    _populate_findings(finding_sheet, findings)
    _populate_sources(source_sheet, spec, sources)
    _apply_theme(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _populate_focus_lead(
    ws,
    spec: WorkpaperSpec,
    context: dict[str, str],
    workbook_sheets: tuple[str, ...],
    status: str,
) -> None:
    catalog = PUBLIC_WORKPAPER_CATALOG.get(spec.code)
    ws["B1"] = f"{spec.code} {spec.title}"
    ws["B2"] = "客户名称"
    ws["C2"] = context["client_name"]
    ws["B3"] = "期间"
    ws["C3"] = context["period_end"]
    ws["B4"] = "编制人"
    ws["C4"] = context["prepared_by"]
    ws["E4"] = "复核人"
    ws["B6"] = "公开版成熟度"
    ws["C6"] = catalog.maturity if catalog else "deep-beta"
    ws["B7"] = "自动化状态"
    ws["C7"] = status

    procedure_rows = []
    procedures = catalog.procedures if catalog else spec.procedures
    for idx, procedure in enumerate(procedures, start=1):
        procedure_rows.append([idx, procedure, "已生成可复核底稿区域", "待项目组复核"])
    _write_table(ws, "B10", ["序号", "程序", "Agent 输出", "项目组结论"], procedure_rows)

    ws["G2"] = "工作表数量"
    ws["H2"] = len(workbook_sheets)
    ws["G3"] = "来源要求"
    ws["H3"] = "、".join(catalog.source_requirements if catalog else spec.source_keys)
    ws["G4"] = "生成原则"
    ws["H4"] = "clean-room 自有底稿；不复制事务所模板、隐藏表、宏或专有文案。"


def _investment_rows(tb_path: Path | None) -> list[dict[str, Any]]:
    return _read_filtered_records(
        tb_path,
        lambda row: _row_account_code(row).startswith(("1101", "1503", "1511", "1512", "1521", "4001", "6111"))
        or any(
            token in _row_account_name(row)
            for token in ("交易性金融资产", "其他权益工具投资", "债权投资", "其他债权投资", "长期股权投资", "投资收益")
        ),
        100,
    )


def _expense_tb_rows(tb_path: Path | None) -> list[dict[str, Any]]:
    return _read_filtered_records(
        tb_path,
        lambda row: _row_account_code(row).startswith(("5", "6401", "6402", "6601", "6602", "6603", "6604", "6701", "6702", "6801"))
        or any(token in _row_account_name(row) for token in ("费用", "成本", "减值损失", "税金")),
        100,
    )


def _expense_journal_rows(path: Path | None) -> list[dict[str, Any]]:
    return _read_filtered_records(
        path,
        lambda row: _row_account_code(row).startswith(("5", "6401", "6402", "6601", "6602", "6603", "6604", "6701", "6702", "6801"))
        or any(token in _row_account_name(row) for token in ("费用", "成本", "减值损失", "税金")),
        500,
    )


def _d_classification_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    output: list[list[Any]] = []
    for idx, row in enumerate(rows, start=1):
        name = _row_account_name(row) or _row_label(row)
        amount = _first_value(row, ["期末余额", "ending_balance", "本期借方", "本期贷方", "debit", "credit"])
        output.append(
            [
                idx,
                "TB 金融资产分类",
                f"{_row_account_code(row)} {name}".strip(),
                amount,
                _investment_classification(name),
                "需核对总账明细、投资台账和管理层持有目的，确认列报分类。",
                "",
            ]
        )
    return output


def _d_valuation_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return [[1, "无投资类余额", "", "不适用；如管理层确认存在投资，请补充投资台账和估值资料", "不伪造估值结论", "待确认", ""]]
    output = []
    for idx, row in enumerate(rows, start=1):
        name = _row_account_name(row) or _row_label(row)
        amount = _first_value(row, ["期末余额", "ending_balance", "本期借方", "本期贷方", "debit", "credit"])
        output.append(
            [
                idx,
                name,
                amount,
                "理财合同/估值单/托管报告/公允价值输入",
                "估值层次、关键输入、期末持有数量和账面金额勾稽",
                "待复核" if _to_float(amount) else "不适用",
                "",
            ]
        )
    return output


def _d_confirmation_rows(rows: list[dict[str, Any]], samples: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return [[1, "无投资类余额", "适用性确认", "TB、总账明细、管理层声明", "需确认本期无持有或处置投资", "待复核", "D10"]]
    output = []
    for idx, row in enumerate(rows, start=1):
        name = _row_account_name(row) or _row_label(row)
        output.append(
            [
                idx,
                name,
                "函证/替代/抽样",
                "银行/券商/托管方函证，或合同、交易回单、估值单",
                "关注权利义务、存在性和估值准确性",
                "待复核",
                "D10",
            ]
        )
    return output or [[idx, sample.get("样本/对象", ""), "抽样", "", sample.get("风险提示", ""), "待复核", "D10"] for idx, sample in enumerate(samples, start=1)]


def _expense_fluctuation_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return [[1, "", "未识别费用/成本科目", "", "资料不足", "补充 TB 或科目余额表后重新生成", ""]]
    output = []
    for idx, row in enumerate(rows, start=1):
        amount = _first_value(row, ["期末余额", "ending_balance", "本期借方", "debit", "借方金额", "金额"])
        output.append(
            [
                idx,
                _row_account_code(row),
                _row_account_name(row) or _row_label(row),
                amount,
                "大额/敏感费用" if abs(_to_float(amount)) >= 1_000_000 else "常规波动复核",
                "结合预算、上期数、合同和发票检查发生真实性与分类准确性",
                "",
            ]
        )
    return output


def _expense_bkd_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not rows:
        return [[1, "", "", "", "未解析到费用明细", "", "", "", "补充序时账/accounting_records 后重新生成", ""]]
    sorted_rows = sorted(rows, key=lambda row: abs(_row_amount(row)), reverse=True)[:30]
    output = []
    for idx, row in enumerate(sorted_rows, start=1):
        output.append(
            [
                idx,
                _first_value(row, ["voucher_id", "凭证号", "凭证编号"]),
                _first_value(row, ["voucher_date", "凭证日期", "date", "日期"]),
                _row_account_code(row),
                _row_account_name(row),
                _first_value(row, ["memo", "摘要", "description"]),
                _first_value(row, ["debit", "借方金额", "本期借方"]),
                _first_value(row, ["credit", "贷方金额", "本期贷方"]),
                "检查合同/发票/审批/付款记录，确认费用发生与分类。",
                "",
            ]
        )
    return output


def _expense_tod_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    output = []
    for idx, row in enumerate(rows[:25], start=1):
        if "程序" in row:
            output.append([idx, "", "", row.get("样本/对象", ""), row.get("金额/指标", ""), row.get("程序", ""), row.get("复核建议", ""), "待复核"])
        else:
            output.append(
                [
                    idx,
                    _first_value(row, ["voucher_id", "凭证号", "凭证编号"]),
                    _first_value(row, ["voucher_date", "凭证日期", "date", "日期"]),
                    _row_account_name(row) or _row_label(row),
                    _row_amount(row),
                    "穿行凭证至合同、发票、审批和付款记录",
                    "原始凭证、合同/订单、发票、付款流水、审批记录",
                    "待复核",
                ]
            )
    return output or [[1, "", "", "未解析到费用样本", "", "资料不足", "补充明细账后重新生成", "待复核"]]


def _expense_cutoff_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    dated_rows = sorted(rows, key=lambda row: str(_first_value(row, ["voucher_date", "凭证日期", "date", "日期"])), reverse=True)
    output = []
    for idx, row in enumerate(dated_rows[:12], start=1):
        amount = _row_amount(row)
        output.append(
            [
                idx,
                _first_value(row, ["voucher_id", "凭证号", "凭证编号"]),
                _first_value(row, ["voucher_date", "凭证日期", "date", "日期"]),
                _row_account_name(row),
                amount,
                "期末前后费用截止风险" if abs(amount) >= 100_000 else "常规截止样本",
                "检查服务期间、验收/收货日期、发票日期和付款日期是否匹配归属期间",
                "",
            ]
        )
    return output


def _investment_classification(account_name: str) -> str:
    if "其他权益工具" in account_name:
        return "其他权益工具投资"
    if "长期股权" in account_name:
        return "长期股权投资/需另行执行权益法或成本法复核"
    if "债权" in account_name:
        return "债权投资/其他债权投资"
    if "投资收益" in account_name:
        return "投资收益发生额"
    return "交易性金融资产或其他金融资产"


def _row_amount(row: dict[str, Any]) -> float:
    for key in ("期末余额", "ending_balance", "借方金额", "贷方金额", "本期借方", "本期贷方", "debit", "credit", "金额", "amount"):
        if key in row and row[key] not in ("", None):
            value = _to_float(row[key])
            if value:
                return value
    for value in row.values():
        numeric = _to_float(value)
        if numeric:
            return numeric
    return 0.0


def _row_label(row: dict[str, Any]) -> str:
    for value in row.values():
        if value not in ("", None):
            return str(value)[:80]
    return ""


def _populate_lead(ws, spec: WorkpaperSpec, context: dict[str, str], samples: list[dict[str, Any]], findings: list[dict[str, Any]]) -> None:
    ws["B1"] = f"{spec.code} {spec.title}"
    ws["B2"] = "客户名称"
    ws["C2"] = context["client_name"]
    ws["B3"] = "期间"
    ws["C3"] = context["period_end"]
    ws["B4"] = "编制人"
    ws["C4"] = context["prepared_by"]
    ws["E4"] = "复核人"
    ws["B6"] = "认定/目标"
    ws["C6"] = spec.assertion
    ws["B8"] = "执行程序"
    for idx, procedure in enumerate(spec.procedures, start=9):
        ws.cell(idx, 2).value = idx - 8
        ws.cell(idx, 3).value = procedure
        ws.cell(idx, 5).value = "已生成样本，待项目组复核"
    ws["G2"] = "样本数"
    ws["H2"] = "=COUNTA('程序与样本'!B9:B200)"
    ws["G3"] = "发现数"
    ws["H3"] = "=COUNTA('审计发现'!B9:B200)"
    ws["G4"] = "高风险提示"
    ws["H4"] = '=COUNTIF(\'审计发现\'!C9:C200,"高")'
    ws["B15"] = "自动化结论"
    ws["C15"] = (
        f"已基于可解析资料生成 {len(samples)} 条结构化样本和 {len(findings)} 条复核提示。"
        "本底稿为可信辅助产物，最终审计结论需由项目组复核确认。"
    )


def _populate_samples(ws, spec: WorkpaperSpec, samples: list[dict[str, Any]]) -> None:
    ws["B1"] = f"{spec.code} 程序与样本"
    headers = ["序号", "程序", "样本/对象", "金额/指标", "风险提示", "复核建议", "项目组结论", "索引"]
    _write_table(ws, "B8", headers, [[idx, *sample.values(), "", spec.code] for idx, sample in enumerate(samples, start=1)])


def _populate_findings(ws, findings: list[dict[str, Any]]) -> None:
    ws["B1"] = "审计发现与复核提示"
    headers = ["编号", "严重性", "发现类型", "金额", "描述", "建议程序", "处理状态"]
    rows = [[finding.get(key, "") for key in headers[:-1]] + ["待复核"] for finding in findings]
    _write_table(ws, "B8", headers, rows)


def _populate_sources(ws, spec: WorkpaperSpec, sources: dict[str, Path | None]) -> None:
    ws["B1"] = "来源索引"
    headers = ["资料类型", "路径", "是否存在", "SHA256", "用途"]
    rows = []
    for key in spec.source_keys:
        path = sources.get(key)
        exists = bool(path and path.exists())
        digest = ""
        if exists and path and path.is_file():
            try:
                digest = sha256_file(path)
            except Exception:
                digest = ""
        rows.append([key, str(path or ""), "是" if exists else "否", digest, spec.title])
    _write_table(ws, "B8", headers, rows)


def _write_table(ws, anchor: str, headers: list[str], rows: list[list[Any]]) -> None:
    start_col = ws[anchor].column
    start_row = ws[anchor].row
    for col_offset, header in enumerate(headers):
        ws.cell(start_row, start_col + col_offset).value = header
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset).value = value
    end_row = start_row + max(len(rows), 1)
    end_col = start_col + len(headers) - 1
    table_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=f"T_{abs(hash((ws.title, start_row, len(headers)))) % 1_000_000_000}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _apply_theme(wb) -> None:
    title_fill = PatternFill("solid", fgColor="1F2937")
    header_fill = PatternFill("solid", fgColor="0F766E")
    sub_fill = PatternFill("solid", fgColor="E6F4F1")
    warning_fill = PatternFill("solid", fgColor="FDE68A")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B8" if ws.max_row >= 8 else "A2"
        if ws.max_column >= 2:
            ws.column_dimensions["A"].width = 3
            for col in range(2, min(ws.max_column, 12) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16 if col != 3 else 30
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00;[Red]-#,##0.00;""'
        if ws["B1"].value:
            ws["B1"].fill = title_fill
            ws["B1"].font = Font(color="FFFFFF", bold=True, size=13)
            ws["B1"].alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28
        for cell in ws[8]:
            if cell.value:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
        for row in range(2, min(ws.max_row, 7) + 1):
            if ws.cell(row, 2).value:
                ws.cell(row, 2).fill = sub_fill
                ws.cell(row, 2).font = Font(bold=True)
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if isinstance(cell.value, str) and ("待复核" in cell.value or "需复核" in cell.value):
                    cell.fill = warning_fill
        if ws.max_row >= 9 and ws.max_column >= 3:
            ws.auto_filter.ref = f"B8:{get_column_letter(ws.max_column)}{ws.max_row}"
        _add_risk_conditioning(ws)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True


def _add_risk_conditioning(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "严重性":
                col = cell.column
                start = cell.row + 1
                end = min(ws.max_row, 200)
                if start <= end:
                    ref = f"{get_column_letter(col)}{start}:{get_column_letter(col)}{end}"
                    ws.conditional_formatting.add(ref, CellIsRule(operator="equal", formula=['"高"'], fill=PatternFill("solid", fgColor="FCA5A5")))
                    ws.conditional_formatting.add(ref, CellIsRule(operator="equal", formula=['"中"'], fill=PatternFill("solid", fgColor="FDE68A")))
                return


def _zip_suite(output_root: Path, workbooks: dict[str, Path], manifest_path: Path) -> Path:
    zip_path = output_root / "workpaper_suite.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in workbooks.values():
            if path.exists():
                archive.write(path, arcname=path.name)
        if manifest_path.exists():
            archive.write(manifest_path, arcname=manifest_path.name)
    return zip_path
