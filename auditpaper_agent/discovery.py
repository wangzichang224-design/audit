from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path


GENERATED_EXCEL_STEMS = {"out", "filled", "filled_workpaper"}
NON_CASH_JOURNAL_TOKENS = (
    "ap_subledger",
    "ar_subledger",
    "receivable",
    "payable",
    "supplier",
    "vendor",
    "customer",
    "应收",
    "应付",
    "供应商",
    "客户",
)


@dataclass(frozen=True)
class CashMaterialsDiscovery:
    materials_dir: Path
    trial_balance: Path | None = None
    journal: Path | None = None
    bank_statement: Path | None = None
    template: Path | None = None
    confidence: float = 0.0
    candidate_sets: list["CashMaterialSet"] = field(default_factory=list)
    agent_used: bool = False
    agent_reason: str = ""

    @property
    def missing_required(self) -> list[str]:
        missing: list[str] = []
        if self.trial_balance is None:
            missing.append("试算平衡表/TB")
        if self.journal is None:
            missing.append("银行日记账/序时账")
        if self.template is None:
            missing.append("底稿模板")
        return missing


@dataclass(frozen=True)
class CashMaterialSet:
    root: Path
    discovery: CashMaterialsDiscovery
    score: int
    confidence: float
    reasons: list[str] = field(default_factory=list)


def discover_cash_materials(materials_dir: str | Path) -> CashMaterialsDiscovery:
    return resolve_cash_materials(materials_dir, use_agent=False)


def resolve_cash_materials(materials_dir: str | Path, use_agent: bool = True) -> CashMaterialsDiscovery:
    root = Path(materials_dir)
    if not root.is_dir():
        raise NotADirectoryError(f"资料文件夹不存在：{root}")

    material_sets = discover_cash_material_sets(root)
    if material_sets:
        if use_agent and _should_ask_agent(material_sets):
            agent_result = _resolve_with_agent(root, material_sets)
            if agent_result is not None:
                return agent_result
        best = material_sets[0]
        return CashMaterialsDiscovery(
            materials_dir=best.discovery.materials_dir,
            trial_balance=best.discovery.trial_balance,
            journal=best.discovery.journal,
            bank_statement=best.discovery.bank_statement,
            template=best.discovery.template,
            confidence=best.confidence,
            candidate_sets=material_sets[:8],
            agent_used=False,
            agent_reason="rule-based coherent material set selection",
        )

    return _discover_in_dir(root)


def discover_cash_material_sets(root: str | Path) -> list[CashMaterialSet]:
    path = Path(root)
    if not path.is_dir():
        return []

    scored: list[CashMaterialSet] = []
    for candidate in _candidate_material_roots(path):
        discovery = _discover_in_dir(candidate)
        score, reasons = _score_material_set(path, candidate, discovery)
        if score <= 0:
            continue
        scored.append(
            CashMaterialSet(
                root=candidate,
                discovery=discovery,
                score=score,
                confidence=round(min(score / 230, 1.0), 3),
                reasons=reasons,
            )
        )
    scored.sort(key=lambda item: (-item.score, len(item.root.parts), str(item.root).lower()))
    return scored


def _discover_in_dir(root: Path) -> CashMaterialsDiscovery:
    files = sorted([p for p in root.rglob("*") if p.is_file()], key=lambda p: (len(p.parts), p.name))
    dirs = sorted([p for p in root.rglob("*") if p.is_dir()], key=lambda p: (len(p.parts), p.name))

    trial_balance = _pick_file(
        files,
        include_any=("试算", "tb", "trial_balance", "trial balance", "科目余额"),
        exclude_any=("bak", "~$", "底稿", "workpaper", "日记账", "序时", "journal", "gl_balance", "gl_balances"),
    ) or _pick_by_excel_kind(files, "trial_balance")
    journal = _pick_file(
        files,
        include_any=("日记账", "序时", "journal", "ledger", "银行存款"),
        exclude_any=("bak", "~$", "底稿", "workpaper", "试算", "trial", *NON_CASH_JOURNAL_TOKENS),
    ) or _pick_by_excel_kind(files, "journal")
    bank_statement = _pick_dir(
        dirs,
        include_any=("bank_confirmations", "confirmations", "回函", "函证", "对账单", "statement"),
    ) or _pick_file(
        files,
        include_any=("回函", "函证", "对账单", "statement", "confirmation"),
        exclude_any=("bak", "~$", "底稿", "workpaper"),
    )
    template = _pick_file(
        files,
        include_any=("底稿", "workpaper", "template", "货币资金"),
        exclude_any=("bak", "~$", "试算", "日记账", "序时", "journal", "ledger", "对账单", "回函", "函证"),
    )

    return CashMaterialsDiscovery(
        materials_dir=root,
        trial_balance=trial_balance,
        journal=journal,
        bank_statement=bank_statement,
        template=template,
    )


def _candidate_material_roots(root: Path) -> list[Path]:
    candidates: dict[str, Path] = {str(root.resolve()).lower(): root}
    for path in root.rglob("*"):
        if not path.is_dir():
            continue
        if _looks_like_cash_material_dir(path):
            candidates[str(path.resolve()).lower()] = path
        if (path / "c底稿资料").is_dir():
            candidates[str(path.resolve()).lower()] = path
            candidates[str((path / "c底稿资料").resolve()).lower()] = path / "c底稿资料"
        if (path / "04_accounting_records").is_dir() and (path / "05_audit_workpapers").is_dir():
            candidates[str(path.resolve()).lower()] = path
    return sorted(candidates.values(), key=lambda p: (len(p.parts), str(p).lower()))


def _score_material_set(input_root: Path, candidate: Path, discovery: CashMaterialsDiscovery) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []
    if discovery.trial_balance:
        score += 45
        reasons.append("trial balance detected")
    if discovery.journal:
        score += 45
        reasons.append("cash journal detected")
    if discovery.template:
        score += 45
        reasons.append("cash template detected")
    if discovery.bank_statement:
        score += 25
        reasons.append("bank statement/confirmation detected")
    if _looks_like_cash_material_dir(candidate):
        score += 25
        reasons.append("cash material directory")
    if (candidate / "c底稿资料").is_dir():
        score += 15
        reasons.append("project root with cash materials")
    if discovery.journal and _looks_like_non_cash_subledger(discovery.journal):
        score -= 80
        reasons.append("journal candidate looks like AP/AR subledger")
    if discovery.journal:
        score += min(_inspect_table_score(discovery.journal, "journal"), 20)
    if discovery.trial_balance:
        score += min(_inspect_table_score(discovery.trial_balance, "trial_balance"), 20)
    try:
        if candidate == input_root and len(list(input_root.iterdir())) > 12:
            score -= 20
            reasons.append("large parent folder fallback")
    except Exception:
        pass
    if discovery.missing_required:
        score -= 35 * len(discovery.missing_required)
        reasons.append("missing required: " + ", ".join(discovery.missing_required))
    return score, reasons


def _should_ask_agent(material_sets: list[CashMaterialSet]) -> bool:
    if not material_sets:
        return True
    best = material_sets[0]
    if best.confidence >= 0.8 and not best.discovery.missing_required and _looks_like_cash_material_dir(best.root):
        return False
    if best.confidence < 0.72 or best.discovery.missing_required:
        return True
    if len(material_sets) > 1 and best.score - material_sets[1].score <= 12:
        return True
    return False


def _resolve_with_agent(root: Path, material_sets: list[CashMaterialSet]) -> CashMaterialsDiscovery | None:
    try:
        from auditpaper_agent.agent import AgentReasoningProvider

        result = AgentReasoningProvider().select_cash_materials(root, material_sets)
    except Exception:
        return None
    if not result.selected_paths:
        return None

    selected = result.selected_paths
    discovery = CashMaterialsDiscovery(
        materials_dir=Path(selected.get("materials_dir") or root),
        trial_balance=Path(selected["trial_balance"]) if selected.get("trial_balance") else None,
        journal=Path(selected["journal"]) if selected.get("journal") else None,
        bank_statement=Path(selected["bank_statement"]) if selected.get("bank_statement") else None,
        template=Path(selected["template"]) if selected.get("template") else None,
        confidence=result.confidence,
        candidate_sets=material_sets[:8],
        agent_used=True,
        agent_reason=result.reason,
    )
    if any(path and not path.exists() for path in [discovery.trial_balance, discovery.journal, discovery.template]):
        return None
    return discovery


def _pick_file(files: list[Path], include_any: tuple[str, ...], exclude_any: tuple[str, ...]) -> Path | None:
    candidates = [
        p
        for p in files
        if p.suffix.lower() in {".xlsx", ".xls", ".xlsm", ".csv", ".pdf", ".png", ".jpg", ".jpeg"}
        and not _looks_like_generated_output(p)
    ]
    scored: list[tuple[int, Path]] = []
    for path in candidates:
        haystack = path.name.lower()
        if any(token.lower() in haystack for token in exclude_any):
            continue
        score = sum(1 for token in include_any if token.lower() in haystack)
        if not score:
            continue
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            score += 1
        scored.append((score, path))
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], len(item[1].parts), item[1].name))
    return scored[0][1]


def _pick_by_excel_kind(files: list[Path], kind: str) -> Path | None:
    scored: list[tuple[int, Path]] = []
    for path in files:
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls", ".csv"}:
            continue
        if path.name.startswith("~$"):
            continue
        if _looks_like_generated_output(path):
            continue
        if kind == "journal" and _looks_like_non_cash_subledger(path):
            continue
        score = _inspect_table_score(path, kind)
        if score:
            scored.append((score, path))
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], len(item[1].parts), item[1].name))
    return scored[0][1]


def _inspect_table_score(path: Path, kind: str) -> int:
    try:
        if path.suffix.lower() == ".csv":
            import pandas as pd

            cols = [str(c) for c in pd.read_csv(path, encoding="utf-8-sig", nrows=0).columns]
            sheet_names: list[str] = []
        else:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet_names = list(wb.sheetnames)
                ws = wb[wb.sheetnames[0]]
                cols = []
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
                    cols.extend(str(v) for v in row if v is not None)
            finally:
                wb.close()
    except Exception:
        return 0

    text = " ".join([path.name, *sheet_names, *cols]).lower()
    if kind == "trial_balance":
        keys = (
            "科目编码",
            "科目名称",
            "期末借方",
            "期末贷方",
            "上年末审定",
            "trial",
            "account_code",
            "account_name",
        )
        negative = ("日期", "摘要", "对方单位", "交易", "银行名称", "账号")
    else:
        keys = (
            "日期",
            "摘要",
            "借方金额",
            "贷方金额",
            "余额",
            "银行名称",
            "账号",
            "对方单位",
            "voucher",
            "journal",
        )
        negative = ("科目编码", "科目名称", "期末借方", "期末贷方", *NON_CASH_JOURNAL_TOKENS)
    score = sum(1 for key in keys if key.lower() in text)
    score -= sum(1 for key in negative if key.lower() in text)
    return max(score, 0)


def _pick_dir(dirs: list[Path], include_any: tuple[str, ...]) -> Path | None:
    scored: list[tuple[int, Path]] = []
    for path in dirs:
        haystack = str(path).lower()
        score = sum(1 for token in include_any if token.lower() in haystack)
        if score:
            scored.append((score, path))
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], len(item[1].parts), item[1].name))
    return scored[0][1]


def _looks_like_cash_material_dir(path: Path) -> bool:
    text = str(path).lower()
    return any(token in text for token in ("c底稿资料", "货币资金", "cash_case", "cash materials", "cash_materials"))


def _looks_like_generated_output(path: Path) -> bool:
    """Avoid treating prior agent outputs as source materials."""
    return path.stem.lower() in GENERATED_EXCEL_STEMS


def _looks_like_non_cash_subledger(path: Path) -> bool:
    text = str(path).lower()
    return any(token.lower() in text for token in NON_CASH_JOURNAL_TOKENS)
