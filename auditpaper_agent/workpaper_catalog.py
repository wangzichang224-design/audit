from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROHIBITED_TEMPLATE_MARKERS = (
    "PROPRIETARY_SETTING_SHEET",
    "PRIVATE_INTERNAL",
    "THIRD_PARTY_TEMPLATE_MARKER",
)


@dataclass(frozen=True)
class WorkpaperCatalogItem:
    code: str
    title: str
    maturity: str
    source_requirements: tuple[str, ...]
    procedures: tuple[str, ...]
    output_sheets: tuple[str, ...]
    public_note: str


@dataclass(frozen=True)
class TemplateSheetSummary:
    name: str
    state: str
    risky_marker: bool = False


@dataclass(frozen=True)
class TemplateFileSummary:
    path: Path
    file_name: str
    suffix: str
    mapped_code: str
    sheet_count: int
    hidden_sheets: list[str] = field(default_factory=list)
    risky_sheets: list[str] = field(default_factory=list)
    sheets: list[TemplateSheetSummary] = field(default_factory=list)
    has_macro: bool = False
    readable: bool = True
    error: str = ""


@dataclass(frozen=True)
class TemplateInventorySummary:
    reference_root: Path
    template_count: int
    supplementary_count: int
    readable_count: int
    mapped_subjects: dict[str, int]
    files: list[TemplateFileSummary]
    clean_room_warnings: list[str]
    public_catalog_codes: list[str]

    @property
    def has_prohibited_markers(self) -> bool:
        return any(file.risky_sheets or file.has_macro for file in self.files)


PUBLIC_WORKPAPER_CATALOG: dict[str, WorkpaperCatalogItem] = {
    "C": WorkpaperCatalogItem(
        code="C",
        title="货币资金",
        maturity="deep-automated",
        source_requirements=("trial_balance", "bank_journal", "bank_statement_or_confirmation", "cash_template"),
        procedures=(
            "Lead 波动和 TB tie-out",
            "银行账户全量 BKD",
            "函证/对账单跟进",
            "特殊账户和受限资金分类检查",
            "银行调节、外币测算、截止测试",
            "负数现金和现金盘点复核提示",
        ),
        output_sheets=(
            "汇总",
            "C.00 Lead",
            "C.00 BKD",
            "C.01 Confirmations",
            "C.01b 特殊账户程序",
            "C.00b 分类检查",
            "C.02 Bank reconciliations",
            "C.02a 外币测算",
            "C.03 Cutoff",
        ),
        public_note="自有现金底稿生成器，保留公式保护和 provenance 批注。",
    ),
    "D10": WorkpaperCatalogItem(
        code="D10",
        title="交易性金融资产及其他权益工具投资",
        maturity="deep-beta",
        source_requirements=("trial_balance", "accounting_records", "investment_listing_optional"),
        procedures=(
            "金融资产适用性和分类检查",
            "公允价值/理财/结构性存款样本复核",
            "函证与抽样结果跟进",
            "无余额时输出不适用判断底稿",
        ),
        output_sheets=("Lead", "金融资产分类", "公允价值复核", "函证与抽样结果", "审计发现", "来源索引"),
        public_note="公开版只做可复核结构化判断，不自动声称完成投资类审计结论。",
    ),
    "EXP10": WorkpaperCatalogItem(
        code="EXP10",
        title="费用测试与波动分析",
        maturity="deep-beta",
        source_requirements=("trial_balance", "accounting_records", "purchase_cycle_optional"),
        procedures=(
            "费用科目波动分析",
            "费用明细 BKD 和大额样本",
            "TOD 详细测试和敏感费用复核",
            "期末截止测试提示",
        ),
        output_sheets=("Lead", "费用波动", "明细BKD", "TOD详细测试", "截止测试", "审计发现", "来源索引"),
        public_note="公开版提供抽样和复核提示，最终审计判断由项目组确认。",
    ),
    "A10": WorkpaperCatalogItem(
        code="A10",
        title="风险评估与重要性",
        maturity="scaffold-ready",
        source_requirements=("trial_balance", "tie_out", "validation_report"),
        procedures=("TB tie-out", "重要性和重大波动识别", "后续程序关注点"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿，不替代完整风险评估。",
    ),
    "E20": WorkpaperCatalogItem(
        code="E20",
        title="应收账款与 ECL",
        maturity="scaffold-ready",
        source_requirements=("customers", "sales_cycle", "trial_balance"),
        procedures=("客户样本", "账龄/ECL 初步提示", "人工复核结论位"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿，不自动完成 ECL 专家判断。",
    ),
    "F10": WorkpaperCatalogItem(
        code="F10",
        title="存货监盘与跌价",
        maturity="scaffold-ready",
        source_requirements=("products", "raw_materials", "bom", "inventory_cycle"),
        procedures=("监盘样本", "跌价提示", "来源索引"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿。",
    ),
    "K10": WorkpaperCatalogItem(
        code="K10",
        title="固定资产与折旧",
        maturity="scaffold-ready",
        source_requirements=("fixed_assets", "trial_balance"),
        procedures=("新增/处置样本", "折旧复核提示", "来源索引"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿。",
    ),
    "N10": WorkpaperCatalogItem(
        code="N10",
        title="应付账款与 SURL",
        maturity="scaffold-ready",
        source_requirements=("suppliers", "purchase_cycle", "trial_balance"),
        procedures=("供应商样本", "期后付款/SURL 提示", "来源索引"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿。",
    ),
    "U10": WorkpaperCatalogItem(
        code="U10",
        title="收入确认与截止",
        maturity="scaffold-ready",
        source_requirements=("customers", "sales_cycle", "sales_books", "trial_balance"),
        procedures=("收入样本", "截止测试提示", "来源索引"),
        output_sheets=("Lead", "程序与样本", "审计发现", "来源索引"),
        public_note="第一版为结构化拟真底稿。",
    ),
}


def inspect_template_inventory(reference_root: str | Path) -> TemplateInventorySummary:
    root = _clean_path(reference_root)
    files: list[TemplateFileSummary] = []
    if not root.exists():
        return TemplateInventorySummary(
            reference_root=root,
            template_count=0,
            supplementary_count=0,
            readable_count=0,
            mapped_subjects={},
            files=[],
            clean_room_warnings=[f"Reference template root does not exist: {root}"],
            public_catalog_codes=list(PUBLIC_WORKPAPER_CATALOG),
        )

    for path in sorted(root.rglob("*")):
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        files.append(_inspect_template_file(path, root))

    mapped: dict[str, int] = {}
    for file in files:
        mapped[file.mapped_code] = mapped.get(file.mapped_code, 0) + 1

    warnings = [
        "Clean-room rule: use this inventory for information architecture only; do not copy workbook binaries, hidden sheets, macros, proprietary wording, UUIDs, or styling.",
    ]
    risky = [file.file_name for file in files if file.risky_sheets or file.has_macro]
    if risky:
        warnings.append("Reference files contain hidden/proprietary-risk structures: " + ", ".join(risky[:8]))

    return TemplateInventorySummary(
        reference_root=root,
        template_count=len([file for file in files if "Supplementary" not in file.path.parts]),
        supplementary_count=len([file for file in files if "Supplementary" in file.path.parts]),
        readable_count=len([file for file in files if file.readable]),
        mapped_subjects=mapped,
        files=files,
        clean_room_warnings=warnings,
        public_catalog_codes=list(PUBLIC_WORKPAPER_CATALOG),
    )


def assert_workbook_clean_room(path: str | Path) -> None:
    """Reject generated workbooks that visibly carry proprietary template residue."""
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        for sheet_name in wb.sheetnames:
            if _has_prohibited_marker(sheet_name):
                raise ValueError(f"Generated workbook contains prohibited sheet marker: {sheet_name}")
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 20)):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and _has_prohibited_marker(value):
                        raise ValueError(
                            f"Generated workbook contains prohibited text marker in {sheet_name}!{cell.coordinate}"
                        )
    finally:
        wb.close()


def _inspect_template_file(path: Path, root: Path) -> TemplateFileSummary:
    try:
        wb = load_workbook(path, read_only=False, keep_vba=path.suffix.lower() == ".xlsm", data_only=False)
        try:
            sheets = [
                TemplateSheetSummary(
                    name=ws.title,
                    state=ws.sheet_state,
                    risky_marker=_has_prohibited_marker(ws.title),
                )
                for ws in wb.worksheets
            ]
            hidden = [sheet.name for sheet in sheets if sheet.state != "visible"]
            risky = [sheet.name for sheet in sheets if sheet.risky_marker]
        finally:
            wb.close()
        return TemplateFileSummary(
            path=path,
            file_name=path.name,
            suffix=path.suffix.lower(),
            mapped_code=_infer_subject_code(path.name),
            sheet_count=len(sheets),
            hidden_sheets=hidden,
            risky_sheets=risky,
            sheets=sheets,
            has_macro=path.suffix.lower() == ".xlsm",
        )
    except Exception as exc:
        return TemplateFileSummary(
            path=path,
            file_name=path.name,
            suffix=path.suffix.lower(),
            mapped_code=_infer_subject_code(path.name),
            sheet_count=0,
            has_macro=path.suffix.lower() == ".xlsm",
            readable=False,
            error=str(exc),
        )


def _infer_subject_code(name: str) -> str:
    upper = name.upper()
    if upper.startswith(("C ", "C.")) or "NEGATIVE CASH" in upper or "货币资金" in name:
        return "C"
    if upper.startswith("D ") or "交易性金融资产" in name:
        return "D10"
    if "U_EXP" in upper or "VC&VD" in upper or "费用" in name:
        return "EXP10"
    if upper.startswith("K") or "固定资产" in name:
        return "K10"
    if upper.startswith("N") or "应付账款" in name or "SURL" in upper:
        return "N10"
    if "ECL" in upper or "应收" in name:
        return "E20"
    if upper.startswith("T"):
        return "T"
    if upper.startswith("H"):
        return "H"
    if upper.startswith("J"):
        return "J"
    if upper.startswith("L"):
        return "L"
    return "UNMAPPED"


def _has_prohibited_marker(value: Any) -> bool:
    text = str(value)
    return any(marker in text for marker in PROHIBITED_TEMPLATE_MARKERS)


def _clean_path(value: str | Path) -> Path:
    text = str(value).strip().strip("\ufeff").strip()
    text = text.strip("`").strip().strip('"').strip("'").strip()
    if text.lower().startswith("file://"):
        text = text[7:]
    return Path(text).expanduser()
