from __future__ import annotations

import json
import shutil
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from auditpaper_agent.diagnostics import diagnose_client_package, run_workpaper_stress_suite
from auditpaper_agent.suite import inspect_workpaper_project, run_auto_workpaper_suite
from auditpaper_agent.workpaper_catalog import assert_workbook_clean_room, inspect_template_inventory


def test_inspect_workpaper_project_resolves_cash_folder_to_parent_project() -> None:
    project = _write_project()

    summary = inspect_workpaper_project(project / "c底稿资料")

    assert summary.is_ready
    assert summary.project_dir == project
    assert summary.cash_materials_dir == project / "c底稿资料"
    assert summary.client_name == "华衡汽车部件苏州有限公司"
    assert len(summary.expected_workbooks) == 9
    assert summary.source_coverage["tb"]
    assert summary.source_coverage["master_data"]

    parent_summary = inspect_workpaper_project(project.parent)
    assert parent_summary.is_ready
    assert parent_summary.project_dir == project


def test_run_auto_workpaper_suite_generates_enterprise_workbooks() -> None:
    project = _write_project()
    output_dir = project / "enterprise_out"

    result = run_auto_workpaper_suite(project, output_dir=output_dir)

    assert result.success
    assert result.output_dir == output_dir
    assert result.manifest_path and result.manifest_path.exists()
    assert result.zip_path and result.zip_path.exists()
    assert set(result.workbooks) == {"A10", "C", "D10", "E20", "EXP10", "F10", "K10", "N10", "U10"}

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["schema_version"] == "auditpaper.workpaper_suite.v1"
    assert manifest["client_name"] == "华衡汽车部件苏州有限公司"
    assert len(manifest["workbooks"]) == 9

    cash_wb = load_workbook(result.workbooks["C"], data_only=False)
    try:
        assert {"汇总", "C.00 Lead", "C.00 BKD", "C.01 Confirmations", "C.02 Bank reconciliations", "C.03 Cutoff"}.issubset(
            set(cash_wb.sheetnames)
        )
        assert not any("PRIVATE_INTERNAL" in name or name == "PROPRIETARY_SETTING_SHEET" for name in cash_wb.sheetnames)
        assert cash_wb["C.00 Lead"]["C2"].value == "华衡汽车部件苏州有限公司"
        assert cash_wb["C.00 Lead"]["D29"].value == "=SUM(D9:D28)"
        assert "SHA256" in cash_wb["C.00 Lead"]["C2"].comment.text
        assert cash_wb["汇总"]["J1"].value == "公开版复核摘要"
        assert cash_wb["汇总"]["K4"].value and "COUNTIF" in cash_wb["汇总"]["K4"].value
    finally:
        cash_wb.close()

    generic_wb = load_workbook(result.workbooks["E20"], data_only=False)
    try:
        assert generic_wb.sheetnames == ["Lead", "程序与样本", "审计发现", "来源索引"]
        assert generic_wb["Lead"]["C2"].value == "华衡汽车部件苏州有限公司"
        assert generic_wb["Lead"]["H2"].value == "=COUNTA('程序与样本'!B9:B200)"
        assert generic_wb["审计发现"]["B9"].value == "E20-F01"
    finally:
        generic_wb.close()

    exp_wb = load_workbook(result.workbooks["EXP10"], data_only=False)
    try:
        assert exp_wb.sheetnames == ["Lead", "费用波动", "明细BKD", "TOD详细测试", "截止测试", "审计发现", "来源索引"]
        assert exp_wb["Lead"]["C2"].value == "华衡汽车部件苏州有限公司"
        assert exp_wb["费用波动"]["D9"].value == "管理费用"
        assert exp_wb["明细BKD"]["C9"].value == "V-EXP-001"
        assert exp_wb["TOD详细测试"]["G9"].value == "穿行凭证至合同、发票、审批和付款记录"
    finally:
        exp_wb.close()

    d_wb = load_workbook(result.workbooks["D10"], data_only=False)
    try:
        assert d_wb.sheetnames == ["Lead", "金融资产分类", "公允价值复核", "函证与抽样结果", "审计发现", "来源索引"]
        assert "未识别交易性金融资产" in d_wb["金融资产分类"]["D9"].value
        assert "不适用判断" in d_wb["Lead"]["C7"].value
    finally:
        d_wb.close()

    for workbook in result.workbooks.values():
        assert_workbook_clean_room(workbook)


def test_template_inventory_and_clean_room_checks() -> None:
    root = _runtime_dir() / "reference"
    supplementary = root / "Supplementary"
    supplementary.mkdir(parents=True)

    risky = root / "C Reference 货币资金 202YMMDD XYZ公司.xlsx"
    wb = Workbook()
    wb.active.title = "汇总"
    hidden = wb.create_sheet("PROPRIETARY_SETTING_SHEET")
    hidden.sheet_state = "hidden"
    wb.save(risky)
    wb.close()

    macro = root / "H 自研参考 202YMMDD XYZ公司.xlsm"
    macro.write_bytes(b"not a real workbook")
    _write_rows(supplementary / "C.00a Negative cash.xlsx", "negative_cash", [["item"]])

    inventory = inspect_template_inventory(root)

    assert inventory.template_count == 2
    assert inventory.supplementary_count == 1
    assert inventory.readable_count == 2
    assert inventory.mapped_subjects["C"] == 2
    assert inventory.has_prohibited_markers
    assert "C" in inventory.public_catalog_codes

    good = _runtime_dir() / "good.xlsx"
    _write_rows(good, "Lead", [["自有底稿"]])
    assert_workbook_clean_room(good)

    bad = _runtime_dir() / "bad.xlsx"
    _write_rows(bad, "PRIVATE_INTERNAL", [["hidden marker"]])
    with pytest.raises(ValueError, match="prohibited"):
        assert_workbook_clean_room(bad)


def test_diagnose_client_package_modes() -> None:
    project = _write_project()
    full = diagnose_client_package(project)

    assert full.mode == "full_project"
    assert full.can_generate
    assert full.provider_summary["api_default"].startswith("offline rules first")
    assert full.selected_cash_materials["journal"].endswith("序时账_2025.xlsx")
    assert any(check.name == "provider_default_offline" and check.passed for check in full.checks)

    cash_only = _runtime_dir() / "cash_only"
    shutil.copytree(project / "c底稿资料", cash_only)
    cash_diag = diagnose_client_package(cash_only)

    assert cash_diag.mode == "single_cash"
    assert cash_diag.can_generate

    invalid = _runtime_dir() / "invalid"
    invalid.mkdir()
    invalid_diag = diagnose_client_package(invalid)

    assert invalid_diag.mode == "diagnostic_only"
    assert not invalid_diag.can_generate
    assert invalid_diag.missing_required


def test_workpaper_stress_suite_handles_valid_and_invalid_cases() -> None:
    fixtures = _runtime_dir() / "fixtures"
    fixtures.mkdir()
    shutil.copytree(_write_project(), fixtures / "valid_project")
    (fixtures / "invalid_project").mkdir()

    result = run_workpaper_stress_suite(fixtures, focus=("C", "D", "EXP"))

    assert result.case_count == 2
    assert result.runnable_count == 1
    assert result.success_count == 1
    assert result.invalid_handled_count == 1
    assert result.success_rate == 1.0


def _runtime_dir() -> Path:
    path = Path(__file__).resolve().parents[1] / "test_runtime" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_project() -> Path:
    root = _runtime_dir() / "audit_sim_autoparts_2025_v2"
    cash = root / "c底稿资料"
    statements = cash / "银行对账单"
    statements.mkdir(parents=True)
    (root / "04_accounting_records").mkdir(parents=True)
    (root / "02_master_data").mkdir(parents=True)
    (root / "03_business_cycles" / "sales").mkdir(parents=True)
    (root / "03_business_cycles" / "purchase").mkdir(parents=True)
    (root / "03_business_cycles" / "production_inventory").mkdir(parents=True)
    (root / "03_business_cycles" / "payroll_tax_fixed_assets").mkdir(parents=True)
    (root / "03_business_cycles" / "monthly_document_books" / "sales").mkdir(parents=True)
    (root / "05_audit_workpapers").mkdir(parents=True)
    (root / "06_validation").mkdir(parents=True)

    _write_rows(
        cash / "试算平衡表_TB_2025.xlsx",
        "TB",
        [
            ["科目编码", "科目名称", "期初余额", "本期借方", "本期贷方", "期末余额", "余额方向"],
            ["1002", "银行存款", 100, 0, 0, 11448673.71, "借"],
        ],
    )
    _write_rows(
        root / "04_accounting_records" / "TB_2025.xlsx",
        "TB",
        [
            ["科目编码", "科目名称", "期末余额"],
            ["1002", "银行存款", 11448673.71],
            ["1122", "应收账款", 2200000],
            ["6602", "管理费用", 1200000],
        ],
    )
    _write_rows(
        cash / "序时账_2025.xlsx",
        "序时账",
        [
            ["凭证日期", "凭证号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额", "来源单号"],
            ["2025-12-31", "V-001", "银行收款", "1002", "银行存款", 100, 0, "BANK-001"],
        ],
    )
    _write_rows(
        root / "04_accounting_records" / "序时账_2025.xlsx",
        "序时账",
        [
            ["凭证日期", "凭证号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额"],
            ["2025-12-31", "V-001", "银行收款", "1002", "银行存款", 100, 0],
        ],
    )
    _write_rows(
        root / "04_accounting_records" / "accounting_records.xlsx",
        "records",
        [
            ["voucher_id", "voucher_date", "account_code", "account_name", "debit", "credit", "memo"],
            ["V-EXP-001", "2025-12-20", "6602", "管理费用", 1200000, 0, "管理费用样本"],
        ],
    )
    _write_rows(
        statements / "银行对账单_BA_CMB_02_2025.xlsx",
        "招商银行苏州分行",
        [
            ["交易序号", "交易日期", "流水号", "摘要", "借方发生额", "贷方发生额", "余额", "凭证号", "对账状态"],
            [1, "2025-12-31", "BANK-001", "期末余额", 0, 0, 11448673.71, "V-001", "已达账"],
        ],
    )
    _write_cash_template(cash / "C_货币资金_空白底稿模板.xlsx")
    (cash / "企业信用报告_华衡汽车部件苏州有限公司.pdf").write_bytes(b"%PDF-1.4\n")

    _write_csv(root / "02_master_data" / "customers.csv", "customer_id,customer_name,credit_limit\nC001,客户一,1200000\n")
    _write_csv(root / "02_master_data" / "suppliers.csv", "supplier_id,supplier_name,credit_limit\nS001,供应商一,900000\n")
    _write_csv(root / "02_master_data" / "products.csv", "product_id,product_name,standard_cost\nP001,控制臂,350\n")
    _write_csv(root / "02_master_data" / "raw_materials.csv", "material_id,material_name,unit_cost\nR001,钢材,80\n")
    _write_csv(root / "02_master_data" / "bom.csv", "product_id,material_id,qty\nP001,R001,2\n")
    _write_csv(root / "02_master_data" / "fixed_assets_master.csv", "asset_id,asset_name,cost\nFA001,冲压设备,1500000\n")
    (root / "02_master_data" / "master_data.xlsx").write_bytes(b"placeholder")

    _write_rows(root / "03_business_cycles" / "sales" / "sales_cycle.xlsx", "sales", [["order_id", "amount"], ["SO-001", 1600000]])
    _write_rows(root / "03_business_cycles" / "purchase" / "purchase_cycle.xlsx", "purchase", [["po_id", "amount"], ["PO-001", 800000]])
    _write_rows(root / "03_business_cycles" / "production_inventory" / "production_inventory_cycle.xlsx", "inventory", [["sku", "amount"], ["P001", 500000]])
    _write_rows(
        root / "03_business_cycles" / "payroll_tax_fixed_assets" / "payroll_tax_fixed_assets.xlsx",
        "fixed_assets",
        [["asset_id", "cost"], ["FA001", 1500000]],
    )
    (root / "03_business_cycles" / "monthly_document_books" / "sales" / "sales_book_12.pdf").write_bytes(b"%PDF-1.4\n")
    _write_rows(root / "06_validation" / "tie_out_matrix.xlsx", "tie", [["item", "status"], ["TB", "OK"]])
    (root / "06_validation" / "validation_report.json").write_text('{"status":"OK"}', encoding="utf-8")
    return root


def _write_csv(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8-sig")


def _write_rows(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_cash_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "汇总": [["索引", "工作内容", "执行", "不执行原因", "备注"]],
        "C.00 Lead": [["科目编码", "科目名称", "TB期末数", "A3审前数", "银行对账单余额", "未达调节", "调节后账面数", "差异", "结论", "资料索引"]],
        "C.00 BKD": [["公司名称", "科目", "账户ID", "开户行", "账号", "账户性质", "币种", "期末账面余额", "对账单余额", "是否取得对账单", "是否纳入函证", "备注"]],
        "C.01 Confirmations": [["账户ID", "开户行", "账号", "账面余额", "函证金额", "发函日期", "回函日期", "回函金额", "差异", "差异说明", "替代程序", "结论"]],
        "C.01b 特殊账户程序": [["检查项目", "检查方法", "项目组填写结果", "是否需进一步程序", "结论"]],
        "C.00b 分类检查": [["账户ID", "开户行", "账面余额", "是否受限", "受限原因", "列报分类", "检查证据", "结论"]],
        "C.02 Bank reconciliations": [["账户ID", "开户行", "银行对账单余额", "加：企业已收银行未收", "减：银行已收企业未入账", "加：企业已付银行未付", "减：银行已付企业未入账", "调节后账面余额", "账面余额", "差异", "结论"]],
        "C.02a 外币测算": [["账户ID", "币种", "原币金额", "汇率", "本币金额", "账面金额", "差异", "结论"]],
        "C.03 Cutoff": [["流水号", "日期", "账户ID", "摘要", "借方", "贷方", "凭证号", "来源单号", "序时账日期", "是否跨期", "结论"]],
    }
    for name, header_rows in sheets.items():
        ws = wb.create_sheet(name)
        if name == "汇总":
            ws["B1"] = "索引"
            ws["C1"] = "工作内容"
            for idx, label in enumerate(["C.00", "C.00 BKD", "C.01", "C.01b", "C.00b", "C.02", "C.02a", "C.03"], start=2):
                ws.cell(idx, 2).value = label
                ws.cell(idx, 3).value = label
            continue
        ws["B1"] = name
        ws["B2"] = "客户名称"
        ws["B3"] = "期末"
        ws["B4"] = "编制人"
        for col, value in enumerate(header_rows[0], start=2):
            ws.cell(8, col).value = value
    wb.save(path)
    wb.close()
