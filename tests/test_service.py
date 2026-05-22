from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from auditpaper_agent.service import inspect_cash_materials, run_auto_cash_case


def test_run_auto_cash_case_from_complete_folder() -> None:
    root = _runtime_dir()
    materials = _write_materials(root / "华衡 货币资金资料")
    case_dir = root / "case"

    result = run_auto_cash_case(f'  "{materials}"  ', case_dir=case_dir)

    assert result.success
    assert result.client_name == "华衡汽车部件苏州有限公司"
    assert result.period_end == "2025-12-31"
    assert result.findings_count >= 0
    assert result.write_commands_count > 0
    assert result.provenance_count == result.write_commands_count
    assert result.output_path and result.output_path.exists()
    assert result.artifacts["case_package"].exists()
    assert result.artifacts["write_plan"].exists()
    assert result.artifacts["provenance"].exists()
    assert {event.stage for event in result.trace_events} >= {"Sense", "Logic", "Harness", "Output"}

    wb = load_workbook(result.output_path)
    try:
        assert wb["C.00 Lead"]["C2"].value == "华衡汽车部件苏州有限公司"
        assert "SHA256" in wb["C.00 Lead"]["C2"].comment.text
    finally:
        wb.close()


def test_inspect_cash_materials_reports_missing_template() -> None:
    root = _runtime_dir()
    materials = _write_materials(root / "缺模板资料", include_template=False)

    discovery = inspect_cash_materials(materials)

    assert not discovery.is_ready
    assert "底稿模板" in discovery.missing_required
    assert discovery.trial_balance is not None
    assert discovery.journal is not None


def test_run_auto_cash_case_does_not_run_when_required_file_missing() -> None:
    root = _runtime_dir()
    materials = _write_materials(root / "缺模板资料", include_template=False)
    case_dir = root / "case"

    result = run_auto_cash_case(materials, case_dir=case_dir)

    assert not result.success
    assert not case_dir.exists()
    assert result.output_path is None
    assert "底稿模板" in result.discovery.missing_required


def test_run_auto_cash_case_returns_error_when_parse_fails() -> None:
    root = _runtime_dir()
    materials = root / "bad_materials"
    materials.mkdir()
    for name in ["试算平衡表.xlsx", "银行存款日记账.xlsx", "C Reference 货币资金 202YMMDD XYZ公司.xlsx"]:
        (materials / name).write_text("not a workbook", encoding="utf-8")

    result = run_auto_cash_case(materials, case_dir=root / "case")

    assert not result.success
    assert result.case_dir == root / "case"
    assert result.output_path is None
    assert "资料解析或底稿生成失败" in result.error


def _runtime_dir() -> Path:
    path = Path(__file__).resolve().parents[1] / "test_runtime" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_materials(root: Path, include_template: bool = True) -> Path:
    statements = root / "银行对账单"
    statements.mkdir(parents=True)
    (root / "企业信用报告_华衡汽车部件苏州有限公司.pdf").write_bytes(b"%PDF-1.4\n")

    _write_rows(
        root / "试算平衡表_TB_2025.xlsx",
        "TB",
        [
            ["科目编码", "科目名称", "期初余额", "本期借方", "本期贷方", "期末余额", "余额方向"],
            ["1002", "银行存款", 100, 0, 0, 11448673.71, "借"],
        ],
    )
    _write_rows(
        root / "序时账_2025.xlsx",
        "序时账",
        [
            ["凭证日期", "凭证号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额", "来源单号"],
            ["2025-12-31", "V-001", "银行收款", "1002", "银行存款", 100, 0, "BANK-001"],
        ],
    )
    _write_rows(
        statements / "银行对账单_BA_CMB_02_2025.xlsx",
        "招商银行苏州分行",
        [
            ["交易序号", "交易日期", "流水号", "摘要", "借方发生额", "贷方发生额", "余额", "凭证号", "对账状态"],
            [1, "2025-12-31", "BANK-001", "期末余额", 0, 0, 11448673.71, "V-001", "已达账"],
        ],
    )
    if include_template:
        _write_sim_template(root / "C_货币资金_空白底稿模板.xlsx")
    return root


def _write_rows(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_sim_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    headers = {
        "C.00 Lead": {
            8: ["科目编码", "科目名称", "TB期末数", "A3审前数", "银行对账单余额", "未达调节", "调节后账面数", "差异", "结论", "资料索引"],
            18: ["账户ID", "开户行", "账号", "账户性质", "币种", "账面余额", "对账单余额", "函证状态", "备注"],
        },
        "C.00 BKD": {
            8: ["公司名称", "科目", "账户ID", "开户行", "账号", "账户性质", "币种", "期末账面余额", "对账单余额", "是否取得对账单", "是否纳入函证"],
        },
        "C.01 Confirmations": {
            8: ["账户ID", "开户行", "账号", "账面余额", "函证金额", "发函日期", "回函日期", "回函金额", "差异", "差异说明", "替代程序"],
        },
        "C.02 Bank reconciliations": {
            8: ["账户ID", "开户行", "银行对账单余额", "加：企业已收银行未收", "减：银行已收企业未入账", "加：企业已付银行未付", "减：银行已付企业未入账", "调节后账面余额", "账面余额", "差异", "结论"],
            25: ["未达编号", "账户ID", "日期", "类型", "金额", "摘要", "期后处理", "结论"],
        },
        "C.03 Cutoff": {
            8: ["流水号", "日期", "账户ID", "摘要", "借方", "贷方", "凭证号", "来源单号", "序时账日期", "是否跨期", "结论"],
        },
    }
    for name, sheet_headers in headers.items():
        ws = wb.create_sheet(name)
        ws["B1"] = name
        ws["B2"] = "客户名称"
        ws["B3"] = "期末"
        ws["B4"] = "编制人"
        for row_num, labels in sheet_headers.items():
            for col_offset, value in enumerate(labels, start=2):
                ws.cell(row_num, col_offset).value = value
    wb.save(path)
    wb.close()
