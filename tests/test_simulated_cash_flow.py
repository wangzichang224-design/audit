from __future__ import annotations

from datetime import date
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from auditpaper_agent.contracts import BankJournalRow, CaseMetadata, SourceRef, StandardizedAuditPackage, TrialBalanceRow
from auditpaper_agent.knowledge.cash import reference_cash_profile
from auditpaper_agent.logic.cash import build_cash_write_plan, run_cash_workflow
from auditpaper_agent.sensing.cash import ingest_cash_case


def test_ingest_understands_simulated_cash_materials() -> None:
    tmp_path = _runtime_dir()
    materials = tmp_path / "c底稿资料"
    statements = materials / "银行对账单"
    statements.mkdir(parents=True)
    (materials / "企业信用报告_华衡汽车部件苏州有限公司.pdf").write_bytes(b"%PDF-1.4\n")

    tb = materials / "试算平衡表_TB_2025.xlsx"
    _write_rows(
        tb,
        "TB",
        [
            ["科目编码", "科目名称", "期初余额", "本期借方", "本期贷方", "期末余额", "余额方向"],
            ["1002", "银行存款", 220000000, 82697818.76, 236427945.96, 66269872.8, "借"],
            ["1122", "应收账款", 156000000, 0, 0, 215117663.95, "借"],
        ],
    )
    journal = materials / "序时账_2025.xlsx"
    _write_rows(
        journal,
        "序时账",
        [
            ["序号", "凭证日期", "凭证号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额", "来源类型", "来源单号"],
            [1, "2025-12-31", "V-001", "银行收款", "1002", "银行存款", 100, 0, "BANK", "BANK-001"],
            [2, "2025-12-31", "V-002", "销售收入", "6001", "主营业务收入", 0, 100, "BANK", "BANK-001"],
        ],
    )
    _write_rows(
        statements / "银行对账单_BA_CMB_02_2025.xlsx",
        "招商银行苏州分行",
        [
            ["交易序号", "交易日期", "流水号", "摘要", "借方发生额", "贷方发生额", "余额", "凭证号", "对账状态"],
            [1, "2025-12-29", "BANK-001", "销售回款", 555102.05, 0, 11448673.71, "V-001", "已达账"],
            [2, "2025-12-31", "BANK-002", "银行已收企业未入账-客户尾款", 1680000, 0, 13128673.71, None, "银行已收企业未入账"],
        ],
    )

    package = ingest_cash_case(tmp_path / "case", tb, journal, statements, ocr_provider="pdf-text")

    assert package.meta.client_name == "华衡汽车部件苏州有限公司"
    assert package.meta.period_end.isoformat() == "2025-12-31"
    assert package.trial_balance[0].ending_balance == 66269872.8
    assert len(package.bank_journal) == 1
    assert len(package.bank_confirmations) == 1
    assert package.bank_confirmations[0].bank_name == "招商银行苏州分行"
    assert package.bank_confirmations[0].confirmed_balance == 13128673.71
    assert len(package.bank_statement_transactions) == 2


def test_simulated_cash_template_writes_visible_cells() -> None:
    tmp_path = _runtime_dir()
    materials = tmp_path / "materials"
    statements = materials / "银行对账单"
    statements.mkdir(parents=True)
    tb = materials / "试算平衡表_TB_2025.xlsx"
    journal = materials / "序时账_2025.xlsx"
    template = materials / "C_货币资金_空白底稿模板.xlsx"
    output = tmp_path / "filled.xlsx"

    _write_rows(
        tb,
        "TB",
        [
            ["科目编码", "科目名称", "期初余额", "本期借方", "本期贷方", "期末余额", "余额方向"],
            ["1002", "银行存款", 100, 0, 0, 11448673.71, "借"],
        ],
    )
    _write_rows(
        journal,
        "序时账",
        [
            ["凭证日期", "凭证号", "摘要", "科目编码", "科目名称", "借方金额", "贷方金额", "来源单号"],
            ["2025-12-31", "V-001", "银行收款", "1002", "银行存款", 100, 0, "BANK-001"],
        ],
    )
    _write_rows(
        statements / "银行对账单_BA_CMB_02_2025.xlsx",
        "招商银行苏州分行",
        [
            ["交易序号", "交易日期", "流水号", "摘要", "借方发生额", "贷方发生额", "余额", "凭证号", "对账状态"],
            [1, "2025-12-31", "BANK-001", "期末余额", 0, 0, 11448673.71, "V-001", "已达账"],
        ],
    )
    _write_sim_template(template)

    case_dir = tmp_path / "case"
    ingest_cash_case(case_dir, tb, journal, statements, client_name="测试客户", ocr_provider="pdf-text")
    run_cash_workflow(case_dir, template_path=template, output_path=output)

    wb = load_workbook(output, data_only=False)
    try:
        lead = wb["C.00 Lead"]
        assert lead["C2"].value == "测试客户"
        assert lead["B9"].value == "1002"
        assert lead["D9"].value == 11448673.71
        assert lead["F9"].value == 11448673.71
        assert lead["B38"].value is None
        assert lead["C8"].value == "科目名称"
        assert wb["C.02 Bank reconciliations"]["L9"].value == "调节一致"
    finally:
        wb.close()


def test_reference_cutoff_falls_back_to_near_period_end_large_cash_samples() -> None:
    source = SourceRef(
        document_id="银行存款日记账",
        document_hash="hash",
        file_path="bank_journal.xlsx",
        sheet_name="银行存款日记账",
        row_number=2,
        locator="bank_journal",
    )
    package = StandardizedAuditPackage(
        meta=CaseMetadata(case_id="case", client_name="蓝海跨境", period_end=date(2025, 12, 31)),
        trial_balance=[
            TrialBalanceRow(
                account_code="1002",
                account_name="银行存款",
                ending_debit=11731518.99,
                source=source,
            )
        ],
        bank_journal=[
            BankJournalRow(
                txn_date=date(2025, 12, 27),
                bank_name="工商银行深圳分行",
                bank_account="556633442211009",
                currency="CNY",
                description="汇兑损益-美元结汇",
                debit=5981299.46,
                counterparty="香港子公司",
                source=source,
            ),
            BankJournalRow(
                txn_date=date(2025, 12, 26),
                bank_name="招商银行深圳分行",
                bank_account="755912345678901",
                currency="USD",
                description="跨境平台销售结算款",
                credit=56230.44,
                counterparty="深圳前海供应链有限公司",
                source=source,
            ),
        ],
    )

    plan = build_cash_write_plan(package, [], reference_cash_profile())
    values = {(command.sheet_name, command.cell): command.value for command in plan.commands}

    assert values[("C.03 Cutoff", "C10")].startswith("未识别出明确银行间转账")
    assert values[("C.03 Cutoff", "B20")] == "CUTOFF-01"
    assert values[("C.03 Cutoff", "D20")] == "工商银行深圳分行"
    assert values[("C.03 Cutoff", "I20")] == 5981299.46
    assert "未识别配对银行间转账" in values[("C.03 Cutoff", "N20")]


def _write_rows(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


def _runtime_dir() -> Path:
    path = Path(__file__).resolve().parents[1] / "test_runtime" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_sim_template(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    headers = {
        "C.00 Lead": {
            8: ["科目编码", "科目名称", "TB期末数", "A3审前数", "银行对账单余额", "未达调节", "调节后账面数", "差异", "结论", "资料索引"],
            18: ["账户ID", "开户行", "账号", "账户性质", "币种", "账面余额", "对账单余额", "函证状态", "备注"],
        },
        "C.00 BKD": {
            8: ["公司名称", "科目", "账户ID", "开户行", "账号", "账户性质", "币种", "期末账面余额", "对账单余额", "是否取得对账单", "是否纳入函证"],
        },
        "C.01 Confirmations": {
            8: ["账户ID", "开户行", "账号", "账面余额", "函证金额", "发函日期", "回函日期", "回函金额", "差异", "差异说明", "替代程序"],
        },
        "C.02 Bank reconciliations": {
            8: ["账户ID", "开户行", "银行对账单余额", "加：企业已收银行未收", "减：银行已收企业未入账", "加：企业已付银行未付", "减：银行已付企业未入账", "调节后账面余额", "账面余额", "差异", "结论"],
            25: ["未达编号", "账户ID", "日期", "类型", "金额", "摘要", "期后处理", "结论"],
        },
        "C.03 Cutoff": {
            8: ["流水号", "日期", "账户ID", "摘要", "借方", "贷方", "凭证号", "来源单号", "序时账日期", "是否跨期", "结论"],
        },
    }
    for name, sheet_headers in headers.items():
        ws = wb.create_sheet(name)
        ws["B1"] = name
        ws["B2"] = "客户名称"
        ws["B3"] = "期末"
        ws["B4"] = "编制人"
        ws["B5"] = "资料来源"
        ws["C5"] = "TB / A3 / 序时账 / 银行对账单 / 企业信用报告"
        ws["B6"] = "填表说明"
        ws["C6"] = "空白模板"
        for row_num, labels in sheet_headers.items():
            for col_offset, value in enumerate(labels, start=2):
                ws.cell(row_num, col_offset).value = value
    wb.save(path)
