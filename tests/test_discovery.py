from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

from auditpaper_agent.discovery import discover_cash_materials


def _runtime_dir() -> Path:
    path = Path(__file__).resolve().parents[1] / "test_runtime" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def test_discover_cash_materials_from_one_folder() -> None:
    root = _runtime_dir()
    (root / "bank_confirmations").mkdir()
    tb = root / "试算平衡表.xlsx"
    journal = root / "银行存款日记账.xlsx"
    template = root / "C Reference 货币资金 202YMMDD XYZ公司.xlsx"
    for path in [tb, journal, template, root / "bank_confirmations" / "工商银行_123_询证函回函.pdf"]:
        path.write_text("x", encoding="utf-8")

    result = discover_cash_materials(root)

    assert result.trial_balance == tb
    assert result.journal == journal
    assert result.bank_statement == root / "bank_confirmations"
    assert result.template == template
    assert result.missing_required == []


def test_discover_cash_materials_inside_folder_named_workpaper() -> None:
    root = _runtime_dir() / "c底稿资料"
    root.mkdir()
    (root / "银行对账单").mkdir()
    tb = root / "试算平衡表_TB_2025.xlsx"
    journal = root / "序时账_2025.xlsx"
    template = root / "C_货币资金_空白底稿模板.xlsx"
    for path in [tb, journal, template, root / "企业信用报告.pdf"]:
        path.write_text("x", encoding="utf-8")

    result = discover_cash_materials(root)

    assert result.trial_balance == tb
    assert result.journal == journal
    assert result.bank_statement == root / "银行对账单"
    assert result.template == template


def test_discover_cash_materials_does_not_treat_tb_as_bank_statement() -> None:
    root = _runtime_dir()
    tb = root / "试算平衡表.xlsx"
    journal = root / "银行存款日记账.xlsx"
    template = root / "C Reference 货币资金 202YMMDD XYZ公司.xlsx"
    for path in [tb, journal, template]:
        path.write_text("x", encoding="utf-8")

    result = discover_cash_materials(root)

    assert result.trial_balance == tb
    assert result.journal == journal
    assert result.bank_statement is None
    assert result.template == template
    assert result.missing_required == []


def test_discover_cash_materials_ignores_generated_workbooks() -> None:
    root = _runtime_dir()
    generated = root / "filled_workpaper.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "C.00 Lead"
    ws["B1"] = "标准工作底稿: C Reference 货币资金"
    ws["B2"] = "client"
    wb.create_sheet("C.00 BKD")
    wb.create_sheet("C.01 Confirmations ")
    wb.create_sheet("C.02 Bank reconciliations")
    wb.create_sheet("C.03 Cutoff")
    wb.save(generated)
    wb.close()

    result = discover_cash_materials(root)

    assert result.trial_balance is None
    assert result.journal is None
    assert result.bank_statement is None
    assert result.template is None


def test_discover_cash_materials_prefers_coherent_cash_package_over_parent_noise() -> None:
    root = _runtime_dir()
    cash = root / "audit_sim_autoparts_2025_v2" / "c底稿资料"
    old_data = root / "audit_sim_autoparts_2025" / "02_financial_data"
    v3_confirmations = root / "audit_sim_autoparts_2025_v3_full_pbc" / "02_subject_pbc" / "函证包"
    (cash / "银行对账单").mkdir(parents=True)
    old_data.mkdir(parents=True)
    v3_confirmations.mkdir(parents=True)

    tb = cash / "试算平衡表_TB_2025.xlsx"
    journal = cash / "序时账_2025.xlsx"
    template = cash / "C_货币资金_空白底稿模板.xlsx"
    _write_rows(tb, "TB", [["科目编码", "科目名称", "期末余额"], ["1002", "银行存款", 100]])
    _write_rows(journal, "序时账", [["凭证日期", "摘要", "科目编码", "借方金额", "贷方金额"], ["2025-12-31", "收款", "1002", 100, 0]])
    _write_rows(template, "C.00 Lead", [["底稿"]])
    _write_rows(cash / "银行对账单" / "银行对账单.xlsx", "bank", [["交易日期", "摘要", "余额"], ["2025-12-31", "余额", 100]])
    _write_rows(old_data / "ap_subledger.xlsx", "AP", [["date", "supplier", "amount"], ["2025-12-31", "供应商", 100]])
    (v3_confirmations / "其他项目函证.pdf").write_text("x", encoding="utf-8")

    result = discover_cash_materials(root)

    assert result.materials_dir == cash
    assert result.trial_balance == tb
    assert result.journal == journal
    assert "ap_subledger" not in str(result.journal)
    assert result.template == template
    assert result.confidence > 0.8


def _write_rows(path: Path, sheet_name: str, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
