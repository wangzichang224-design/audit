from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from auditpaper_agent.contracts import SourceRef, WriteCellCommand, WritePlan
from auditpaper_agent.harness.excel import apply_write_plan
from auditpaper_agent.trace import AuditTracer


def _runtime_dir() -> Path:
    path = Path(__file__).resolve().parents[1] / "test_runtime" / uuid4().hex
    path.mkdir(parents=True, exist_ok=True)
    return path


def _source(root: Path) -> SourceRef:
    src = root / "source.txt"
    src.write_text("evidence", encoding="utf-8")
    return SourceRef(
        document_id="source",
        document_hash="hash",
        file_path=str(src),
        sheet_name="Sheet1",
        row_number=1,
    )


def _template(root: Path) -> Path:
    path = root / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "C.00 Lead"
    ws["C2"] = ""
    ws["C3"] = "=TODAY()"
    wb.save(path)
    wb.close()
    return path


def test_harness_writes_allowed_cell_with_comment() -> None:
    root = _runtime_dir()
    template = _template(root)
    plan = WritePlan(
        template_profile="test",
        allowed_cells={"C.00 Lead": ["C2"]},
        commands=[
            WriteCellCommand(
                sheet_name="C.00 Lead",
                cell="C2",
                value="Client A",
                purpose="客户名称",
                source=_source(root),
            )
        ],
    )
    output = root / "out.xlsx"
    result = apply_write_plan(template, output, plan)

    wb = load_workbook(output)
    try:
        assert result.commands_applied == 1
        assert wb["C.00 Lead"]["C2"].value == "Client A"
        assert "SHA256" in wb["C.00 Lead"]["C2"].comment.text
    finally:
        wb.close()


def test_harness_rejects_formula_cell() -> None:
    root = _runtime_dir()
    template = _template(root)
    plan = WritePlan(
        template_profile="test",
        allowed_cells={"C.00 Lead": ["C3"]},
        commands=[
            WriteCellCommand(
                sheet_name="C.00 Lead",
                cell="C3",
                value="2025-12-31",
                purpose="期末",
                source=_source(root),
            )
        ],
    )
    with pytest.raises(ValueError, match="formula"):
        apply_write_plan(template, root / "out.xlsx", plan)


def test_harness_rejects_outside_whitelist() -> None:
    root = _runtime_dir()
    template = _template(root)
    plan = WritePlan(
        template_profile="test",
        allowed_cells={"C.00 Lead": ["C2"]},
        commands=[
            WriteCellCommand(
                sheet_name="C.00 Lead",
                cell="D2",
                value="bad",
                purpose="outside",
                source=_source(root),
            )
        ],
    )
    with pytest.raises(ValueError, match="whitelist"):
        apply_write_plan(template, root / "out.xlsx", plan)


def test_harness_trace_records_write_steps() -> None:
    root = _runtime_dir()
    template = _template(root)
    tracer = AuditTracer(enabled=False)
    plan = WritePlan(
        template_profile="test",
        allowed_cells={"C.00 Lead": ["C2"]},
        commands=[
            WriteCellCommand(
                sheet_name="C.00 Lead",
                cell="C2",
                value="Client A",
                purpose="客户名称",
                source=_source(root),
            )
        ],
    )
    result = apply_write_plan(template, root / "out.xlsx", plan, tracer=tracer)

    assert result.commands_applied == 1
    messages = [event.message for event in tracer.events if event.stage == "Harness"]
    assert "write requested" in messages
    assert "whitelist check passed" in messages
    assert "formula check passed" in messages
    assert "cell written with provenance comment" in messages
